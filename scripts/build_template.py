#!/usr/bin/env python3
"""
EvalFP — Motor de Evaluación Relacional para FP
build_template.py · v2.0-alpha — Multi-módulo, multi-grupo
Generador genérico: los datos del módulo viven en scripts/modules/<modulo>_data.py

Uso:
    python build_template.py               # genera EvalFP.xlsx
    python build_template.py --ia          # genera EvalFP.xlsx + contenido IA (rúbricas, actividades)
    python build_template.py --apuntes     # genera EvalFP.xlsx + apuntes HTML por UT
    python build_template.py --ia --apuntes  # todo a la vez
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
import os
import sys

# ── Cargar datos del módulo ──────────────────────────────────────────────────
# Cambia esta línea para generar el cuaderno de otro módulo.
sys.path.insert(0, os.path.dirname(__file__))
from modules.iso_data import (
    MODULO, UTS, RAS, ASIGNACIONES, EVAL_RAS,
    DUAL_RA, DUAL_PCT_NOTA, DUAL_JUSTIFICACION, RA_INSTRUMENTOS,
)

# ═══════════════════════════════════════════════════════════════════════════════
# PALETA DE COLORES
# ═══════════════════════════════════════════════════════════════════════════════
NAVY   = "1B3A5C"   # Azul marino (cabeceras, nav)
MED    = "2D6A9F"   # Azul medio (nav inactivo)
ICE    = "EBF5FE"   # Azul hielo (fondo alterno)
LBLUE  = "BDD7EE"   # Azul claro (subtítulos, chips)
GOLD   = "FFC000"   # Dorado (separadores)
ORNG   = "E87222"   # Naranja (RA2 accent)
GREEN  = "107C10"   # Verde (disponible)
GRLIT  = "E2EFDA"   # Verde claro (fila disponible)
PURP   = "7030A0"   # Morado (RA7)
PURLIT = "EAE0F5"   # Morado claro
RED    = "C00000"   # Rojo (alertas)
WHITE  = "FFFFFF"
LGRAY  = "F2F2F2"
MGRAY  = "D9D9D9"
DGRAY  = "595959"

# Colores por RA (para badges visuales)
RA_COLORS = {
    "RA1": ("4472C4", "DDEEFF"),  # Azul
    "RA2": ("E87222", "FDE9D7"),  # Naranja
    "RA3": ("70AD47", "E2EFDA"),  # Verde
    "RA4": ("2E75B6", "D6E8F5"),  # Azul oscuro
    "RA5": ("7030A0", "EAE0F5"),  # Morado
    "RA6": ("C00000", "F5DCDC"),  # Rojo
    "RA7": ("375623", "E2EFDA"),  # Verde oscuro
    "RA8": ("833C00", "F9E5D7"),  # Marrón
}

# ═══════════════════════════════════════════════════════════════════════════════
# CONTEXTO DE MÓDULO — prefijo de hojas y helpers de referencia cruzada
# ═══════════════════════════════════════════════════════════════════════════════
_SHEET_PREFIX    = ""    # p.ej. "ISO · " cuando hay varios módulos; "" en mono-módulo
_MODULE_NAV_ITEMS = None # nav items del módulo actual; None → usa NAV_ITEMS global

def _sn(name: str) -> str:
    """Sheet Name con prefijo activo. Ej: _sn('Reg. Notas') → 'ISO · Reg. Notas'"""
    return f"{_SHEET_PREFIX}{name}" if _SHEET_PREFIX else name

def _sr(name: str) -> str:
    """Sheet Reference para fórmulas Excel. Ej: _sr('Reg. Notas') → \"'ISO · Reg. Notas'\" """
    return f"'{_sn(name)}'"

# ═══════════════════════════════════════════════════════════════════════════════
# ÍNDICES RÁPIDOS (calculados a partir de los datos importados)
# ═══════════════════════════════════════════════════════════════════════════════
UT_BY_ID  = {u["id"]: u for u in UTS}
RA_BY_ID  = {r["id"]: r for r in RAS}

def nombre_formula(data_row):
    """Apellidos y Nombre desde la hoja Alumnos.
    Alumnos (Sprint 3): cabecera fila 9, datos en filas 10-39 → alumno i en fila 9+i.
    Hojas de notas: cabecera fila 10, datos en filas 11-40 → alumno i en fila 10+i.
    Por tanto: alumno_row_alumnos = data_row - 1.
    (Sprint 4: corregido de data_row-2 a data_row-1 tras el cambio de Sprint 3.)"""
    al_row = data_row - 1
    return (f"=IFERROR({_sr('Alumnos')}!C{al_row}"
            f"&IF({_sr('Alumnos')}!D{al_row}<>\"\",\", \"&{_sr('Alumnos')}!D{al_row},\"\"),\"\")")

def pond_norm_eval(eval_num):
    """Ponderaciones normalizadas a 100% para los RAs de una evaluación."""
    ra_ids = EVAL_RAS[eval_num]
    total  = sum(RA_BY_ID[rid]["pond"] for rid in ra_ids)
    return {rid: RA_BY_ID[rid]["pond"] / total for rid in ra_ids}

def weighted_formula(col_weight_pairs, row_n, round_digits=2):
    """Devuelve fórmula Excel de suma ponderada: =IFERROR(ROUND(c1*w1+c2*w2+..., 2), '')"""
    parts = [f"{col}{row_n}*{w:.6f}" for col, w in col_weight_pairs]
    return f'=IFERROR(ROUND({"+".join(parts)},{round_digits}),"")'

# CEs únicos por RA (unión de todos los UTs)
def _ces_por_ra():
    ces = {}
    for ut_id, ra_id, celist in ASIGNACIONES:
        ces.setdefault(ra_id, set()).update(celist)
    return {k: sorted(v, key=lambda x: int(x[2:])) for k, v in ces.items()}
CES_POR_RA = _ces_por_ra()

# Horas equivalentes por RA (distribuidas proporcionalmente por CEs en cada UT)
def _horas_por_ra():
    horas = {r["id"]: 0.0 for r in RAS}
    for ut_id, ra_id, ces in ASIGNACIONES:
        total_ces_ut = sum(
            len(c) for u, r, c in ASIGNACIONES if u == ut_id
        )
        h_ut = UT_BY_ID[ut_id]["horas"]
        horas[ra_id] += h_ut * len(ces) / total_ces_ut
    return horas
HORAS_RA = _horas_por_ra()

TOTAL_HORAS = sum(u["horas"] for u in UTS)

# RA → evaluación en que se evalúa principalmente
RA_EVAL_MAP = {rid: ev for ev, rids in EVAL_RAS.items() for rid in rids}

# ── Actividades prácticas generadas desde ASIGNACIONES ─────────────────────
# Regla: una actividad por asignación *primaria* (UT y RA comparten la misma eval).
# Las asignaciones secundarias (e.g. RA7 aparece en UT2 cuya eval es 1)
# se omiten aquí pero siguen figurando en la hoja Programación.
def _actividades_modulo():
    acts = []
    idx  = 1
    for ut in UTS:
        ut_ev = ut.get("eval")
        for ut_id, ra_id, ces in ASIGNACIONES:
            if ut_id != ut["id"]:
                continue
            if RA_EVAL_MAP.get(ra_id) != ut_ev:
                continue   # asignación secundaria → ignorar
            acts.append({
                "id":   f"A{idx:03d}",
                "ut":   ut_id,
                "ra":   ra_id,
                "eval": ut_ev,
                "ces":  ces,
                "desc": f"{ut['nombre']} — práctica {ra_id}",
                "tags": ut.get("tags", ""),
            })
            idx += 1
    return acts

ACTIVIDADES = _actividades_modulo()

# Columna de cada RA en Reg. Notas (referencia cruzada desde las hojas Evaluación).
# Calculado dinámicamente para que sea correcto con cualquier número de actividades.
_RN_ACT_START = 4
_RN_EX_START  = _RN_ACT_START + len(ACTIVIDADES)
_RN_RA_START  = _RN_EX_START  + len(EVAL_RAS)   # EVAL_RAS tiene una entrada por evaluación
RN_RA_COL     = {ra["id"]: get_column_letter(_RN_RA_START + i) for i, ra in enumerate(RAS)}

# Columnas de NOTA EVAL y RESULTADO en cada hoja "Evaluación N"
# build_evaluacion layout: UT_START=4, RA_START=4+len(uts_ev), FINAL_COL=RA_START+len(ra_ids_ev)
def _compute_ev_cols():
    nota = {}; res = {}
    for ev in EVAL_RAS:
        uts_ev = [ut for ut in UTS if ut.get("eval") == ev]
        ra_ev  = EVAL_RAS[ev]
        final  = 4 + len(uts_ev) + len(ra_ev)
        nota[ev] = get_column_letter(final)
        res[ev]  = get_column_letter(final + 1)
    return nota, res
_EV_NOTA_COL, _EV_RES_COL = _compute_ev_cols()

# Columna de NOTA FINAL en 1ªORD  (RA_START_COL=4, len(RAS) columnas RA)
_ORD1_FINAL_COL_L = get_column_letter(4 + len(RAS))     # p.ej. "L"
_ORD1_RES_COL_L   = get_column_letter(4 + len(RAS) + 1) # p.ej. "M"

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS DE APUNTES (integración con build_apuntes.py)
# ═══════════════════════════════════════════════════════════════════════════════
def _kebab(text: str) -> str:
    """Convierte texto con acentos a kebab-case ASCII."""
    import re as _re
    for src, dst in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),
                     ("Á","a"),("É","e"),("Í","i"),("Ó","o"),("Ú","u"),
                     ("ñ","n"),("Ñ","n"),("ü","u"),("Ü","u")]:
        text = text.replace(src, dst)
    text = _re.sub(r"[^a-zA-Z0-9]+", "-", text.lower())
    return text.strip("-")

def _apunte_rel_path(ut: dict) -> str:
    """Ruta relativa desde src/EvalFP.xlsx hasta el index.html del apunte de una UT.
    Ej: '../apuntes/04-FP/ISO/1-asir/instalacion-software-libre/index.html'
    """
    curso_k = _kebab(MODULO.get("curso", "1-asir"))
    ut_k    = _kebab(ut["nombre"])
    return f"../apuntes/04-FP/{MODULO['abrev']}/{curso_k}/{ut_k}/index.html"

def _apunte_rel_path_mod(mod, ut: dict) -> str:
    """Igual que _apunte_rel_path pero acepta el módulo explícitamente (para Biblioteca)."""
    import re as _re
    def _kb(t):
        for s, d in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),
                     ("Á","a"),("É","e"),("Í","i"),("Ó","o"),("Ú","u"),
                     ("ñ","n"),("Ñ","n"),("ü","u"),("Ü","u")]:
            t = t.replace(s, d)
        return _re.sub(r"[^a-zA-Z0-9]+", "-", t.lower()).strip("-")
    curso_k = _kb(mod.MODULO.get("curso", "1-asir"))
    ut_k    = _kb(ut["nombre"])
    return f"../apuntes/04-FP/{mod.MODULO['abrev']}/{curso_k}/{ut_k}/index.html"

# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS DE ESTILO
# ═══════════════════════════════════════════════════════════════════════════════
def f(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def fn(name="Calibri", size=11, bold=False, italic=False, color="000000"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=color)

def al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color=MGRAY):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(color=MGRAY):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def thick_border_bottom(color=NAVY):
    s = Side(style="medium", color=color)
    return Border(bottom=s)

def style_cell(cell, fill_color=None, font_color="000000", font_size=11,
               bold=False, italic=False, h_align="left", v_align="center",
               wrap=False, border=None):
    if fill_color:
        cell.fill = f(fill_color)
    cell.font = fn(size=font_size, bold=bold, italic=italic, color=font_color)
    cell.alignment = al(h=h_align, v=v_align, wrap=wrap)
    if border:
        cell.border = border

def set_col_width(ws, col_widths):
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

# ═══════════════════════════════════════════════════════════════════════════════
# SET MODULE CONTEXT — reinicializa todas las constantes globales de módulo
# ═══════════════════════════════════════════════════════════════════════════════
def _set_module_context(mod, prefix=""):
    """Carga un nuevo módulo y recalcula todas las constantes derivadas.
    Llama a esta función antes de generar las hojas de cada módulo."""
    global MODULO, UTS, RAS, ASIGNACIONES, EVAL_RAS
    global DUAL_RA, DUAL_PCT_NOTA, DUAL_JUSTIFICACION, RA_INSTRUMENTOS
    global UT_BY_ID, RA_BY_ID, CES_POR_RA, HORAS_RA, TOTAL_HORAS
    global RA_EVAL_MAP, ACTIVIDADES
    global _RN_ACT_START, _RN_EX_START, _RN_RA_START, RN_RA_COL
    global _EV_NOTA_COL, _EV_RES_COL, _ORD1_FINAL_COL_L, _ORD1_RES_COL_L
    global _SHEET_PREFIX, _MODULE_NAV_ITEMS

    MODULO           = mod.MODULO
    UTS              = mod.UTS
    RAS              = mod.RAS
    ASIGNACIONES     = mod.ASIGNACIONES
    EVAL_RAS         = mod.EVAL_RAS
    DUAL_RA          = mod.DUAL_RA
    DUAL_PCT_NOTA    = mod.DUAL_PCT_NOTA
    DUAL_JUSTIFICACION = mod.DUAL_JUSTIFICACION
    RA_INSTRUMENTOS  = mod.RA_INSTRUMENTOS
    _SHEET_PREFIX    = prefix

    UT_BY_ID    = {u["id"]: u for u in UTS}
    RA_BY_ID    = {r["id"]: r for r in RAS}
    RA_EVAL_MAP = {rid: ev for ev, rids in EVAL_RAS.items() for rid in rids}
    CES_POR_RA  = _ces_por_ra()
    HORAS_RA    = _horas_por_ra()
    TOTAL_HORAS = sum(u["horas"] for u in UTS)
    ACTIVIDADES = _actividades_modulo()

    _RN_ACT_START     = 4
    _RN_EX_START      = _RN_ACT_START + len(ACTIVIDADES)
    _RN_RA_START      = _RN_EX_START  + len(EVAL_RAS)
    RN_RA_COL         = {ra["id"]: get_column_letter(_RN_RA_START + i) for i, ra in enumerate(RAS)}
    _EV_NOTA_COL, _EV_RES_COL = _compute_ev_cols()
    _ORD1_FINAL_COL_L = get_column_letter(4 + len(RAS))
    _ORD1_RES_COL_L   = get_column_letter(4 + len(RAS) + 1)

    # Nav items del módulo (con prefijo aplicado a cada hoja destino)
    _MODULE_NAV_ITEMS = [
        ("← Módulos",    "Mis Módulos"),
        ("📋 Programa",  _sn("Programación")),
        ("👥 Alumnos",   _sn("Alumnos")),
        ("📝 Actividades", _sn("Actividades")),
        ("📊 Reg. Notas", _sn("Reg. Notas")),
        ("📈 Resumen",   _sn("Resumen")),
        ("Eval 1",       _sn("Evaluación 1")),
        ("Eval 2",       _sn("Evaluación 2")),
        ("Eval 3",       _sn("Evaluación 3")),
        ("1ª ORD",       _sn("1ªORD")),
        ("2ª ORD",       _sn("2ªORD")),
        ("🎯 Rúbricas",  _sn("Rúbricas")),
        ("📊 Inf.Grupo", _sn("Informe Grupo")),
        ("📄 Boletín",   _sn("Boletín")),
        ("📊 Dashboard", _sn("Dashboard")),
    ]

# ═══════════════════════════════════════════════════════════════════════════════
# NAVEGACIÓN Y CABECERA ESTÁNDAR
# ═══════════════════════════════════════════════════════════════════════════════
# Nav global — hojas sin prefijo de módulo (Inicio, Config, Mis Módulos, Biblioteca)
GLOBAL_NAV_ITEMS = [
    ("🏠 Inicio",      "Inicio"),
    ("⚙️ Config",      "Configuración"),
    ("📚 Módulos",     "Mis Módulos"),
    ("📖 Biblioteca",  "Biblioteca"),
    ("📅 Calendario",  "Calendario"),
    ("☀️ Hoy",         "Hoy"),
    ("📊 Dashboard+",  "Dashboard Global"),
]

# Nav legado (modo mono-módulo sin prefijo) — se mantiene por compatibilidad
NAV_ITEMS = [
    ("🏠 Inicio",       "Inicio"),
    ("📚 Módulos",       "Mis Módulos"),
    ("📋 Programación", "Programación"),
    ("👥 Alumnos",      "Alumnos"),
    ("📝 Actividades",  "Actividades"),
    ("📊 Reg. Notas",   "Reg. Notas"),
    ("📈 Resumen",      "Resumen"),
    ("Eval 1",          "Evaluación 1"),
    ("Eval 2",          "Evaluación 2"),
    ("Eval 3",          "Evaluación 3"),
    ("1ª ORD",          "1ªORD"),
    ("2ª ORD",          "2ªORD"),
    ("🎯 Rúbricas",     "Rúbricas"),
    ("📊 Inf.Grupo",    "Informe Grupo"),
    ("📄 Boletín",      "Boletín"),
    ("📊 Dashboard",    "Dashboard"),
]

def apply_header(ws, title, subtitle=""):
    # Fila 1: barra fina decorativa
    ws.row_dimensions[1].height = 5
    for c in range(1, 30):
        ws.cell(1, c).fill = f(GOLD)

    # Fila 2: título
    ws.row_dimensions[2].height = 26
    for c in range(1, 30):
        ws.cell(2, c).fill = f(NAVY)
    ws.merge_cells("B2:K2")
    hc = ws.cell(2, 2, value=title)
    hc.font = fn(size=20, bold=True, color=WHITE)
    hc.alignment = al(h="left", v="center")

    # Fila 3: subtítulo
    ws.row_dimensions[3].height = 18
    for c in range(1, 30):
        ws.cell(3, c).fill = f(NAVY)
    if subtitle:
        ws.merge_cells("B3:K3")
        sc = ws.cell(3, 2, value=subtitle)
        sc.font = fn(size=10, italic=True, color=LBLUE)
        sc.fill = f(NAVY)
        sc.alignment = al(h="left", v="center")

    # Fila 4: separador dorado
    ws.row_dimensions[4].height = 3
    for c in range(1, 30):
        ws.cell(4, c).fill = f(GOLD)

    # Fila 5: barra de navegación
    ws.row_dimensions[5].height = 22

    # Fila 6: separador claro
    ws.row_dimensions[6].height = 8
    for c in range(1, 30):
        ws.cell(6, c).fill = f(ICE)

def apply_nav(ws, current, items=None):
    """Pinta la barra de navegación en fila 5.
    items: lista de (label, target); si None usa _MODULE_NAV_ITEMS o NAV_ITEMS."""
    nav = items or _MODULE_NAV_ITEMS or NAV_ITEMS
    for c in range(1, 30):
        ws.cell(5, c).fill = f(NAVY)
    col = 2
    for label, target in nav:
        cell = ws.cell(5, col, value=label)
        is_active = (target == current)
        cell.fill = f(NAVY if is_active else MED)
        cell.font = fn(size=9, bold=is_active, color=WHITE)
        cell.alignment = al(h="center", v="center")
        if not is_active:
            cell.value = f'=HYPERLINK("#{target}!A1","{label}")'
        col += 1

def apply_standard_header(ws, title, subtitle, current_sheet):
    apply_header(ws, title, subtitle)
    apply_nav(ws, current_sheet)
    # Freeze panes after row 6
    ws.freeze_panes = "B7"

def apply_global_header(ws, title, subtitle, current_sheet):
    """Cabecera para hojas globales (Inicio, Config, Mis Módulos) — nav global."""
    apply_header(ws, title, subtitle)
    apply_nav(ws, current_sheet, items=GLOBAL_NAV_ITEMS)
    ws.freeze_panes = "B7"

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: INICIO
# ═══════════════════════════════════════════════════════════════════════════════
def build_dashboard_global(wb, modules_config):
    """Hoja global con KPIs consolidados de toda la carga docente.
    Estadísticas por módulo y comparativa del curso completo."""
    ws = wb.create_sheet("Dashboard Global")
    apply_global_header(
        ws,
        "📊 Dashboard Global — Curso Completo",
        f"Estadísticas consolidadas por módulo y grupo · Curso {modules_config[0][0].MODULO['anno']}",
        "Dashboard Global",
    )
    set_col_width(ws, {
        "A":"2","B":"14","C":"30","D":"12","E":"12",
        "F":"12","G":"12","H":"12","I":"14",
    })

    # ── Cabecera KPI por módulo ────────────────────────────────────────────────
    r = 8
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  RESUMEN POR MÓDULO — datos en tiempo real desde cada Dashboard")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 28
    hdrs = [
        (2, "Módulo",       NAVY),
        (3, "Nombre",       NAVY),
        (4, "Activos",      MED),
        (5, "✅ Aptos",     "107C10"),
        (6, "❌ No Aptos",  "C00000"),
        (7, "⚠️ En Riesgo", "E87222"),
        (8, "Media grupo",  MED),
        (9, "% Aprobados",  MED),
    ]
    for ci, label, bg in hdrs:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    totals_rows = []  # para la fila de totales
    for mod, grupo in modules_config:
        r += 1
        ws.row_dimensions[r].height = 24
        abrev = mod.MODULO["abrev"]
        prefix = f"{abrev} · "
        dash   = f"{prefix}Dashboard"
        alum   = f"{prefix}Alumnos"
        ord1   = f"{prefix}1ªORD"
        rn     = f"{prefix}Reg. Notas"

        # RA count for nota final column letter
        nf_col = get_column_letter(4 + len(mod.RAS))       # col L for 8 RAs
        res_col = get_column_letter(4 + len(mod.RAS) + 1)  # col M

        row_data = [
            (f'=HYPERLINK("#{dash}!A1","{abrev}")',                      "center", NAVY,     ICE),
            (mod.MODULO["nombre"][:38],                                  "left",   DGRAY,    WHITE),
            (f"=IFERROR(COUNTA('{alum}'!C10:C39),\"\")",                 "center", NAVY,     ICE),
            (f"=IFERROR(COUNTIF('{ord1}'!{res_col}11:{res_col}40,\"APTO\"),\"\")",    "center", "107C10", "E2EFDA"),
            (f"=IFERROR(COUNTIF('{ord1}'!{res_col}11:{res_col}40,\"NO APTO\"),\"\")", "center", "C00000", "F5DCDC"),
            (f"=IFERROR(COUNTIFS('{ord1}'!{nf_col}11:{nf_col}40,\">0\",'{ord1}'!{nf_col}11:{nf_col}40,\"<4\"),\"\")",
             "center", "E87222", "FDE9D7"),
            (f"=IFERROR(ROUND(AVERAGE('{ord1}'!{nf_col}11:{nf_col}40),2),\"\")", "center", NAVY, ICE),
            (f"=IFERROR(TEXT(COUNTIF('{ord1}'!{res_col}11:{res_col}40,\"APTO\")/COUNTA('{ord1}'!{res_col}11:{res_col}40),\"0%\"),\"\")",
             "center", NAVY, WHITE),
        ]
        totals_rows.append(r)
        for ci_off, (val, align, fcolor, fbg) in enumerate(row_data):
            c = ws.cell(r, 2 + ci_off, value=val)
            c.font = fn(size=10, bold=(ci_off == 0), color=fcolor)
            c.fill = f(fbg); c.alignment = al(h=align, v="center")
            c.border = thin_border()
            c.number_format = "0.00" if ci_off == 6 else "@"

    # Fila totales
    r += 1
    ws.row_dimensions[r].height = 22
    for ci, label, bg in [(2,"TOTAL CURSO",NAVY),(3,"",NAVY)]:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg); c.border = thin_border(WHITE)
        c.alignment = al(h="center", v="center")

    if totals_rows:
        first, last = totals_rows[0], totals_rows[-1]
        for ci_off, (agg, fcolor, fbg) in enumerate([
            (f"=SUM(D{first}:D{last})", NAVY,     ICE),
            (f"=SUM(E{first}:E{last})", "107C10", "E2EFDA"),
            (f"=SUM(F{first}:F{last})", "C00000", "F5DCDC"),
            (f"=SUM(G{first}:G{last})", "E87222", "FDE9D7"),
            (f"=IFERROR(ROUND(AVERAGE(H{first}:H{last}),2),\"\")", NAVY, ICE),
            (f"=IFERROR(TEXT(E{r-1}/D{r-1},\"0%\"),\"\")", NAVY, WHITE),
        ]):
            c = ws.cell(r, 4 + ci_off, value=agg)
            c.font = fn(size=10, bold=True, color=fcolor)
            c.fill = f(fbg); c.border = thin_border(WHITE)
            c.alignment = al(h="center", v="center")

    # ── Separador ─────────────────────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 5
    for ci in range(2, 10):
        ws.cell(r, ci).fill = f(GOLD)

    # ── Sección: rendimiento por RA (medias globales entre módulos) ────────────
    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  ACCESOS RÁPIDOS — Dashboards por módulo")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(MED); c.alignment = al(v="center")

    for mod, grupo in modules_config:
        r += 1
        ws.row_dimensions[r].height = 20
        abrev = mod.MODULO["abrev"]
        dash  = f"{abrev} · Dashboard"
        prog  = f"{abrev} · Programación"
        inf   = f"{abrev} · Informe Grupo"

        links = [
            (f'=HYPERLINK("#{dash}!A1","📊 Dashboard {abrev}")', "107C10", "E2EFDA"),
            (f'=HYPERLINK("#{prog}!A1","📋 Programación {abrev}")', MED, ICE),
            (f'=HYPERLINK("#{inf}!A1","📄 Inf.Grupo {abrev}")', MED, ICE),
        ]
        ws.merge_cells(f"B{r}:B{r}")
        ws.cell(r, 2, value=f"{abrev} — {grupo}").font = fn(size=9, bold=True, color=NAVY)
        ws.cell(r, 2).alignment = al(v="center")
        ws.cell(r, 2).border = thin_border()
        for li, (val, fcolor, fbg) in enumerate(links):
            c = ws.cell(r, 3 + li * 2, value=val)
            c.font = fn(size=9, bold=True, color=fcolor)
            c.fill = f(fbg); c.alignment = al(h="center", v="center")
            ws.merge_cells(f"{get_column_letter(3+li*2)}{r}:{get_column_letter(4+li*2)}{r}")
            c.border = thin_border()

    # Pie
    r += 2
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2,
                value="★  Todos los indicadores se calculan automáticamente desde las hojas 1ªORD y Alumnos de cada módulo.")
    c.font = fn(size=8, italic=True, color=DGRAY)
    c.alignment = al(v="center")

    return ws


def build_panel_diario(wb, modules_config):
    """Hoja global 'Hoy' — vista rápida del profesor para el día de trabajo."""
    ws = wb.create_sheet("Hoy")
    apply_global_header(
        ws,
        "☀️ Panel Diario — Hoy",
        "Pendientes · Próximos exámenes · Alumnos en riesgo · Tutorías",
        "Hoy",
    )
    set_col_width(ws, {
        "A":"2","B":"18","C":"35","D":"14","E":"14","F":"14","G":"18",
    })

    anno = modules_config[0][0].MODULO["anno"]

    # ── Bloque 1: Fecha de hoy ─────────────────────────────────────────────────
    r = 8
    ws.row_dimensions[r].height = 30
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2, value='=TEXT(TODAY(),"DDDD, D DE MMMM DE YYYY")')
    c.font = fn(size=16, bold=True, color=NAVY)
    c.alignment = al(h="left", v="center")

    ws.merge_cells(f"E{r}:G{r}")
    c = ws.cell(r, 5, value=f"Curso académico {anno}")
    c.font = fn(size=11, italic=True, color=DGRAY)
    c.alignment = al(h="right", v="center")

    # ── Bloque 2: Módulos activos hoy ─────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 5
    for ci in range(2, 8):
        ws.cell(r, ci).fill = f(GOLD)

    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(r, 2, value="  MIS MÓDULOS HOY")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 20
    for ci, (label, bg) in enumerate([
        ("Módulo", NAVY), ("Nombre completo", NAVY), ("Grupo", MED),
        ("Horas/sem", MED), ("Próx. examen", MED), ("Estado eval", MED),
    ]):
        c = ws.cell(r, 2 + ci, value=label)
        c.font = fn(size=8, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)

    for mod, grupo in modules_config:
        r += 1
        ws.row_dimensions[r].height = 20
        abrev = mod.MODULO["abrev"]
        nombre = mod.MODULO["nombre"]
        horas = mod.MODULO["horas_sem"]
        dash_ref = f"{abrev} · Dashboard"

        row_data = [
            (f'=HYPERLINK("#{dash_ref}!A1","{abrev}")', "center", NAVY, ICE),
            (nombre,  "left",   DGRAY, WHITE),
            (grupo,   "center", MED,   ICE),
            (f"{horas} h/sem", "center", DGRAY, WHITE),
            ("→ ver Calendario", "center", "107C10", "E2EFDA"),
            ("En curso", "center", WHITE, "375623"),
        ]
        for ci_off, (val, align, fcolor, fbg) in enumerate(row_data):
            c = ws.cell(r, 2 + ci_off, value=val)
            c.font = fn(size=9, color=fcolor, bold=(ci_off == 0))
            c.fill = f(fbg); c.alignment = al(h=align, v="center")
            c.border = thin_border()

    # ── Bloque 3: Alumnos en riesgo ────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 5
    for ci in range(2, 8):
        ws.cell(r, ci).fill = f(GOLD)

    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(r, 2, value="  ⚠️ ALUMNOS EN RIESGO — KPIs por módulo (desde Dashboard)")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f("E87222"); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 20
    for ci, (label, bg) in enumerate([
        ("Módulo", NAVY), ("Grupo", MED), ("Activos", MED),
        ("Aptos", "107C10"), ("No Aptos", "C00000"), ("En Riesgo (<4)", "E87222"),
    ]):
        c = ws.cell(r, 2 + ci, value=label)
        c.font = fn(size=8, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)

    for mod, grupo in modules_config:
        r += 1
        ws.row_dimensions[r].height = 20
        abrev = mod.MODULO["abrev"]
        prefix = f"{abrev} · "

        # Referencias dinámicas al Dashboard del módulo
        dash = f"{prefix}Dashboard"
        alum = f"{prefix}Alumnos"

        row_data = [
            (abrev,                                     "center", NAVY,      ICE),
            (grupo,                                     "center", DGRAY,     WHITE),
            (f'=IFERROR(COUNTA(\'{alum}\'!C10:C39),"")', "center", NAVY,    ICE),
            (f'=IFERROR(\'{dash}\'!E10,"–")',            "center", "107C10", "E2EFDA"),
            (f'=IFERROR(\'{dash}\'!G10,"–")',            "center", "C00000", "F5DCDC"),
            (f'=IFERROR(\'{dash}\'!I10,"–")',            "center", "E87222", "FDE9D7"),
        ]
        for ci_off, (val, align, fcolor, fbg) in enumerate(row_data):
            c = ws.cell(r, 2 + ci_off, value=val)
            c.font = fn(size=10, bold=True, color=fcolor)
            c.fill = f(fbg); c.alignment = al(h=align, v="center")
            c.border = thin_border()

    # ── Bloque 4: Tutorías y notas rápidas ────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 5
    for ci in range(2, 8):
        ws.cell(r, ci).fill = f(GOLD)

    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(r, 2, value="  📝 NOTAS RÁPIDAS DEL DÍA — borra y reescribe cada jornada")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(MED); c.alignment = al(v="center")

    for i in range(8):
        r += 1
        ws.row_dimensions[r].height = 20
        bg = ICE if i % 2 == 0 else WHITE
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(r, 2)
        c.fill = f(bg); c.border = thin_border()
        c.font = fn(size=10, color=DGRAY)

    # Pie
    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(r, 2,
                value="★  El bloque 'En Riesgo' lee los KPIs directamente desde el Dashboard de cada módulo.")
    c.font = fn(size=8, italic=True, color=DGRAY)
    c.alignment = al(v="center")

    return ws


def build_calendario(wb, modules_config):
    """Hoja global con el calendario académico del curso completo.
    Permite registrar exámenes, entregas y recuperaciones por módulo y evaluación."""
    ws = wb.create_sheet("Calendario")
    apply_global_header(
        ws,
        "📅 Calendario Académico",
        f"Exámenes · Entregas · Recuperaciones — Curso {modules_config[0][0].MODULO['anno']}",
        "Calendario",
    )
    set_col_width(ws, {
        "A":"2","B":"14","C":"12","D":"12","E":"12",
        "F":"30","G":"20","H":"10","I":"16",
    })

    # ── Leyenda evaluaciones ───────────────────────────────────────────────────
    r = 8
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  CALENDARIO DE EVENTOS ACADÉMICOS — introduce las fechas manualmente")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    # ── Cabecera de tabla ──────────────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 28
    EV_COLORS_CAL = {1: "4472C4", 2: "E87222", 3: "107C10"}
    TYPE_ICONS = {"Examen": "📝", "Entrega": "📤", "Recuperación": "🔄", "Tutoría": "💬", "Otro": "📌"}
    hdrs = [
        (2, "Módulo",      NAVY),
        (3, "Fecha inicio",NAVY),
        (4, "Fecha fin",   NAVY),
        (5, "Evaluación",  NAVY),
        (6, "Descripción", NAVY),
        (7, "Grupo",       MED),
        (8, "Tipo",        MED),
        (9, "Estado",      MED),
    ]
    for ci, label, bg in hdrs:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    ws.freeze_panes = f"B{r+1}"

    # ── 60 filas de eventos (vacías — el profesor las rellena) ────────────────
    TIPOS     = ["Examen", "Entrega", "Recuperación", "Tutoría", "Otro"]
    mod_names = [m.MODULO["abrev"] for m, _ in modules_config]
    grupos    = [g for _, g in modules_config]
    NROWS     = 60
    ROW_START = r + 1       # primera fila de datos
    ROW_END   = r + NROWS   # última fila de datos

    # ── Crear UNA DataValidation por columna (sqref = rango completo) ─────────
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_mod  = DataValidation(type="list", formula1=f'"{",".join(mod_names)}"', allow_blank=True, sqref=f"B{ROW_START}:B{ROW_END}")
    dv_ev   = DataValidation(type="list", formula1='"EV1,EV2,EV3"',            allow_blank=True, sqref=f"E{ROW_START}:E{ROW_END}")
    dv_grp  = DataValidation(type="list", formula1=f'"{",".join(grupos)}"',    allow_blank=True, sqref=f"G{ROW_START}:G{ROW_END}")
    dv_tipo = DataValidation(type="list", formula1=f'"{",".join(TIPOS)}"',     allow_blank=True, sqref=f"H{ROW_START}:H{ROW_END}")
    dv_est  = DataValidation(type="list", formula1='"Pendiente,Confirmado,Realizado,Cancelado"', allow_blank=True, sqref=f"I{ROW_START}:I{ROW_END}")
    for dv in [dv_mod, dv_ev, dv_grp, dv_tipo, dv_est]:
        ws.add_data_validation(dv)

    for i in range(NROWS):
        r += 1
        ws.row_dimensions[r].height = 18
        bg = ICE if i % 2 == 0 else WHITE

        # Col B: Módulo
        c = ws.cell(r, 2)
        c.fill = f(ICE); c.border = thin_border(); c.alignment = al(h="center", v="center")
        c.font = fn(size=9, color=NAVY)

        # Cols C-D: Fechas
        for ci in [3, 4]:
            c = ws.cell(r, ci)
            c.fill = f(bg); c.border = thin_border()
            c.alignment = al(h="center", v="center")
            c.number_format = "DD/MM/YYYY"
            c.font = fn(size=9, color=NAVY)

        # Col E: Evaluación
        c = ws.cell(r, 5)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.font = fn(size=9, color=NAVY)

        # Col F: Descripción
        c = ws.cell(r, 6)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="left", v="center")
        c.font = fn(size=9, color=DGRAY)

        # Col G: Grupo
        c = ws.cell(r, 7)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.font = fn(size=9, color=NAVY)

        # Col H: Tipo
        c = ws.cell(r, 8)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.font = fn(size=9, color=NAVY)

        # Col I: Estado
        c = ws.cell(r, 9)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.font = fn(size=9, color=NAVY)

    # ── Nota pie ──────────────────────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2,
                value="★  Introduce los eventos del curso: exámenes, entregas de proyectos y recuperaciones. "
                      "Usa Ctrl+F para buscar por módulo o grupo.")
    c.font = fn(size=8, italic=True, color=DGRAY)
    c.alignment = al(v="center")

    return ws


def build_biblioteca(wb, modules_config):
    """Hoja global con repositorio de recursos educativos compartidos.
    Incluye actividades, exámenes y rúbricas de todos los módulos activos."""
    ws = wb.create_sheet("Biblioteca")
    apply_global_header(
        ws,
        "📖 Biblioteca de Recursos",
        "Actividades · Exámenes · Rúbricas — compartidos entre módulos",
        "Biblioteca",
    )
    set_col_width(ws, {
        "A":"2","B":"8","C":"10","D":"10","E":"38",
        "F":"12","G":"25","H":"15","I":"8","J":"16",
    })

    # ── Sección: Actividades por módulo ───────────────────────────────────────
    r = 8
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  ACTIVIDADES PRÁCTICAS — índice de todos los módulos")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 26
    hdrs = [
        (2,  "ID",           NAVY),
        (3,  "Módulo",       NAVY),
        (4,  "RA",           NAVY),
        (5,  "Descripción",  NAVY),
        (6,  "Instrumento",  MED),
        (7,  "Tags",         MED),
        (8,  "Eval",         MED),
        (9,  "Peso",         MED),
        (10, "📖 Apuntes",   MED),
    ]
    for ci, label, bg in hdrs:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # Filas de actividades de todos los módulos
    row_count = 0
    for mod_idx, (mod, grupo) in enumerate(modules_config):
        # Recalcular actividades para este módulo
        ev_map   = {rid: ev for ev, rids in mod.EVAL_RAS.items() for rid in rids}
        acts_mod = []
        idx = 1
        for ut in mod.UTS:
            ut_ev = ut.get("eval")
            for ut_id, ra_id, ces in mod.ASIGNACIONES:
                if ut_id != ut["id"]: continue
                if ev_map.get(ra_id) != ut_ev: continue
                acts_mod.append({
                    "id":   f"A{idx:03d}",
                    "ra":   ra_id,
                    "eval": ut_ev,
                    "desc": f"{ut['nombre']} — práctica {ra_id}",
                    "tags": ut.get("tags",""),
                    "peso": "100%",
                    "ut":   ut,       # referencia para hipervínculo apunte
                })
                idx += 1
        # También añadir exámenes (sin UT específica → sin apunte)
        for ev_n, ra_ids in sorted(mod.EVAL_RAS.items()):
            acts_mod.append({
                "id":   f"EX{ev_n}",
                "ra":   ", ".join(ra_ids),
                "eval": ev_n,
                "desc": f"Examen Evaluación {ev_n}",
                "tags": "examen · prueba escrita",
                "peso": "variable",
                "ut":   None,
            })

        abrev = mod.MODULO["abrev"]
        ra_colors_mod = {ra["id"]: RA_COLORS.get(ra["id"], (NAVY, ICE)) for ra in mod.RAS}

        for act in acts_mod:
            r += 1
            row_count += 1
            bg = ICE if row_count % 2 == 0 else WHITE
            ws.row_dimensions[r].height = 18

            ra_id_main = act["ra"].split(",")[0].strip()
            ra_color, ra_lite = ra_colors_mod.get(ra_id_main, (NAVY, ICE))
            is_exam = act["id"].startswith("EX")

            data = [
                (act["id"],                  "center", DGRAY if is_exam else NAVY, bg),
                (abrev,                       "center", WHITE,  ra_color if not is_exam else MED),
                (act["ra"],                   "center", NAVY,   ra_lite if not is_exam else ICE),
                (act["desc"],                 "left",   DGRAY,  bg),
                ("📝 Examen" if is_exam else "🔧 Práctica", "center", DGRAY, bg),
                (act["tags"],                 "left",   DGRAY,  bg),
                (f"EV{act['eval']}",          "center", WHITE,
                 {1:"4472C4",2:"E87222",3:"107C10"}.get(act["eval"], MED)),
                (act["peso"],                 "center", DGRAY,  bg),
            ]
            for ci_off, (val, align, fcolor, fbg) in enumerate(data):
                c = ws.cell(r, 2 + ci_off, value=val)
                c.font = fn(size=9, color=fcolor)
                c.fill = f(fbg); c.alignment = al(h=align, v="center")
                c.border = thin_border()

            # Col J: hipervínculo al apunte de la UT (solo actividades prácticas, no exámenes)
            ut_act = act.get("ut")
            if ut_act and not is_exam:
                apunte_url = _apunte_rel_path_mod(mod, ut_act)
                jc = ws.cell(r, 10, value=f'=HYPERLINK("{apunte_url}","📖 Ver apunte")')
                jc.font = fn(size=9, color="2563EB", italic=True)
                jc.fill = f("EBF5FE")
                jc.alignment = al(h="center", v="center")
                jc.border = thin_border()
            else:
                jc = ws.cell(r, 10, value="")
                jc.fill = f(bg); jc.border = thin_border()

    # ── Separador ─────────────────────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 5
    for ci in range(2, 11):
        ws.cell(r, ci).fill = f(GOLD)

    # ── Sección: Cómo usar la biblioteca ──────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:J{r}")
    c = ws.cell(r, 2, value="  INSTRUCCIONES DE USO")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(MED); c.alignment = al(v="center")

    instrucciones = [
        "CU13 Buscar actividad: usa Ctrl+F en Excel para buscar por módulo, RA o tag.",
        "CU12 Duplicar actividad: copia la fila y pégala en la hoja Actividades del módulo destino.",
        "CU14 Crear rúbrica: ve a la hoja '{ABREV} · Rúbricas' del módulo correspondiente.",
        "CU05 Crear actividad: añade una fila en esta tabla y en la hoja Actividades del módulo.",
        "📖 Apuntes: haz clic en 'Ver apunte' para abrir el HTML de la UT en el navegador.",
        "Para añadir un nuevo módulo edita teacher_config.py y regenera el libro.",
    ]
    for instr in instrucciones:
        r += 1
        ws.row_dimensions[r].height = 16
        ws.merge_cells(f"B{r}:J{r}")
        c = ws.cell(r, 2, value=f"• {instr}")
        c.font = fn(size=9, color=DGRAY)
        c.alignment = al(v="center")

    return ws


def build_mis_modulos(wb, modules_config):
    """Hoja global con la carga docente completa del profesor.
    modules_config: lista de (mod_data, grupo_nombre)."""
    ws = wb.create_sheet("Mis Módulos", 1)   # justo después de Inicio
    apply_global_header(
        ws,
        "📚 Mis Módulos — Carga Docente",
        "Módulos y grupos gestionados en este cuaderno · EvalFP 2.0",
        "Mis Módulos",
    )
    set_col_width(ws, {"A":"2","B":"10","C":"40","D":"15","E":"10",
                        "F":"10","G":"10","H":"18","I":"18"})

    r = 8
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  MÓDULOS ACTIVOS — haz clic en el nombre para ir al módulo")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    # Cabecera tabla
    r += 1
    ws.row_dimensions[r].height = 28
    for ci, (label, bg) in enumerate([
        ("Abrev.", NAVY), ("Módulo", NAVY), ("Grupo", MED),
        ("Horas", MED), ("RAs", MED), ("Evals", MED),
        ("Ir al Inicio", "375623"), ("Ir al Dashboard", "375623"),
    ]):
        c = ws.cell(r, 2 + ci, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # Filas de módulos
    for idx, (mod, grupo) in enumerate(modules_config):
        r += 1
        ws.row_dimensions[r].height = 22
        bg = ICE if idx % 2 == 0 else WHITE
        prefix = f"{mod.MODULO['abrev']} · "

        data = [
            (mod.MODULO["abrev"],   NAVY,  "center"),
            (mod.MODULO["nombre"],  DGRAY, "left"),
            (grupo,                 MED,   "center"),
            (f"{mod.MODULO['total_horas']} h", DGRAY, "center"),
            (str(len(mod.RAS)),     DGRAY, "center"),
            (str(mod.MODULO["eval_count"]),    DGRAY, "center"),
        ]
        for ci, (val, color, align) in enumerate(data):
            c = ws.cell(r, 2 + ci, value=val)
            c.font = fn(size=10, color=color, bold=(ci == 0))
            c.fill = f(bg); c.alignment = al(h=align, v="center")
            c.border = thin_border()

        # Enlace → Programación del módulo
        inicio_sheet = f"{prefix}Programación"
        c = ws.cell(r, 8,
                    value=f'=HYPERLINK("#{inicio_sheet}!A1","→ Programación")')
        c.font = fn(size=9, bold=True, color="107C10")
        c.fill = f("E2EFDA"); c.alignment = al(h="center", v="center")
        c.border = thin_border()

        # Enlace → Dashboard del módulo
        dash_sheet = f"{prefix}Dashboard"
        c = ws.cell(r, 9,
                    value=f'=HYPERLINK("#{dash_sheet}!A1","→ Dashboard")')
        c.font = fn(size=9, bold=True, color="107C10")
        c.fill = f("E2EFDA"); c.alignment = al(h="center", v="center")
        c.border = thin_border()

    # Nota pie
    r += 2
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2,
                value="★  Para añadir un módulo, edita TEACHER_MODULES en scripts/teacher_config.py y regenera el libro.")
    c.font = fn(size=8, italic=True, color=DGRAY)
    c.alignment = al(v="center")

    return ws


def build_inicio(wb):
    ws = wb.create_sheet("Inicio")
    apply_global_header(ws, "EvalFP — Cuaderno del Profesor",
                        "Motor de Evaluación Relacional para Formación Profesional · v2.0",
                        "Inicio")
    set_col_width(ws, {"A":"3","B":"20","C":"25","D":"20","E":"20","F":"20",
                        "G":"20","H":"20","I":"20","J":"20","K":"20"})

    r = 8
    ws.row_dimensions[r].height = 18
    c = ws.cell(r, 2, value="📌  Información del Módulo")
    c.font = fn(size=13, bold=True, color=NAVY)
    c.alignment = al(v="center")

    datos = [
        ("Módulo",      f"{MODULO['nombre']} ({MODULO['abrev']})"),
        ("Código",      MODULO["codigo"]),
        ("Ciclo",       MODULO["ciclo"]),
        ("Curso",       MODULO["curso"]),
        ("Año académico", MODULO["anno"]),
        ("Horas semanales", f"{MODULO['horas_sem']} h/sem"),
        ("Total horas", f"{TOTAL_HORAS} h"),
        ("Evaluaciones",f"{MODULO['eval_count']} evaluaciones"),
    ]
    for i, (label, value) in enumerate(datos):
        r += 1
        is_alt = (i % 2 == 0)
        bg = ICE if is_alt else WHITE
        ws.row_dimensions[r].height = 18
        lc = ws.cell(r, 2, value=label)
        lc.font = fn(bold=True, color=NAVY, size=10)
        lc.fill = f(bg)
        lc.alignment = al(v="center")
        lc.border = thin_border()
        vc = ws.cell(r, 3, value=value)
        vc.font = fn(size=10, color=DGRAY)
        vc.fill = f(bg)
        vc.alignment = al(v="center")
        vc.border = thin_border()

    # Resumen de RAs
    r += 2
    ws.row_dimensions[r].height = 18
    ws.cell(r, 2, value="📊  Ponderación de Resultados de Aprendizaje").font = fn(size=13, bold=True, color=NAVY)

    r += 1
    headers = ["RA", "Resultado de Aprendizaje", "Pond. %", "Horas equiv.", "CEs"]
    for ci, h in enumerate(headers):
        c = ws.cell(r, 2+ci, value=h)
        c.font = fn(bold=True, color=WHITE, size=10)
        c.fill = f(NAVY)
        c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)
    ws.row_dimensions[r].height = 18

    for i, ra in enumerate(RAS):
        r += 1
        is_alt = (i % 2 == 0)
        bg = ICE if is_alt else WHITE
        ra_color, ra_lite = RA_COLORS[ra["id"]]
        ws.row_dimensions[r].height = 18
        cells_data = [
            (ra["id"],             ra_color, WHITE, "center"),
            (ra["nombre"][:80],    bg,        DGRAY, "left"),
            (f'{ra["pond"]}%',     bg,        NAVY,  "center"),
            (f'{HORAS_RA[ra["id"]]:.1f} h', bg, DGRAY, "center"),
            (str(len(CES_POR_RA.get(ra["id"], []))), bg, DGRAY, "center"),
        ]
        for ci, (val, bg_c, fc, ha) in enumerate(cells_data):
            c = ws.cell(r, 2+ci, value=val)
            c.font = fn(size=10, bold=(ci==0), color=fc)
            c.fill = f(bg_c)
            c.alignment = al(h=ha, v="center")
            c.border = thin_border()

    # UTs por evaluación
    r += 2
    ws.cell(r, 2, value="📅  Unidades de Trabajo por Evaluación").font = fn(size=13, bold=True, color=NAVY)
    r += 1
    headers2 = ["UT", "Título", "Horas", "% Total", "Evaluación"]
    for ci, h in enumerate(headers2):
        c = ws.cell(r, 2+ci, value=h)
        c.font = fn(bold=True, color=WHITE, size=10)
        c.fill = f(NAVY)
        c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)
    ws.row_dimensions[r].height = 18

    eval_colors = {"1": "4472C4", "2": "70AD47", "3": "E87222"}
    for i, ut in enumerate(UTS):
        r += 1
        is_alt = (i % 2 == 0)
        bg = ICE if is_alt else WHITE
        pct = ut["horas"] / TOTAL_HORAS * 100
        ev_str = f'Evaluación {ut["eval"]}'
        ev_color = eval_colors[str(ut["eval"])]
        ws.row_dimensions[r].height = 18
        cells_data = [
            (ut["id"],    NAVY,   WHITE,  "center"),
            (ut["nombre"], bg,    DGRAY,  "left"),
            (f'{ut["horas"]}h', bg, DGRAY, "center"),
            (f'{pct:.1f}%', bg, NAVY,    "center"),
            (ev_str,      bg,    ev_color, "center"),
        ]
        for ci, (val, bg_c, fc, ha) in enumerate(cells_data):
            c = ws.cell(r, 2+ci, value=val)
            c.font = fn(size=10, bold=(ci==0), color=fc)
            c.fill = f(bg_c)
            c.alignment = al(h=ha, v="center")
            c.border = thin_border()

    # Total horas
    r += 1
    ws.cell(r, 2, value="TOTAL").font = fn(size=10, bold=True, color=WHITE)
    ws.cell(r, 2).fill = f(NAVY)
    ws.cell(r, 2).border = thin_border()
    ws.cell(r, 4, value=f"{TOTAL_HORAS}h").font = fn(size=10, bold=True, color=WHITE)
    ws.cell(r, 4).fill = f(NAVY)
    ws.cell(r, 4).border = thin_border()
    ws.cell(r, 5, value="100%").font = fn(size=10, bold=True, color=WHITE)
    ws.cell(r, 5).fill = f(NAVY)
    ws.cell(r, 5).border = thin_border()

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: CONFIGURACIÓN
# ═══════════════════════════════════════════════════════════════════════════════
def build_config(wb):
    ws = wb.create_sheet("Configuración")
    apply_global_header(ws, "⚙️ Configuración",
                        "Ajustes del curso, motor de evaluación y preferencias del sistema",
                        "Configuración")
    set_col_width(ws, {"A":"3","B":"28","C":"35","D":"20","E":"20"})

    def section(row, icon, title):
        ws.row_dimensions[row].height = 22
        c = ws.cell(row, 2, value=f"{icon}  {title}")
        c.font = fn(size=12, bold=True, color=WHITE)
        c.fill = f(NAVY)
        c.alignment = al(v="center")
        ws.merge_cells(f"B{row}:E{row}")
        return row

    def field(row, label, value, editable=True):
        ws.row_dimensions[row].height = 18
        lc = ws.cell(row, 2, value=label)
        lc.font = fn(size=10, bold=True, color=NAVY)
        lc.fill = f(LGRAY)
        lc.alignment = al(v="center")
        lc.border = thin_border()
        vc = ws.cell(row, 3, value=value)
        vc.font = fn(size=10, color=DGRAY if not editable else "000000")
        vc.fill = f(WHITE if editable else LGRAY)
        vc.alignment = al(v="center")
        vc.border = thin_border()
        if editable:
            nc = ws.cell(row, 4, value="✏️ editable")
            nc.font = fn(size=9, italic=True, color="70AD47")

    r = 8
    r = section(r, "📚", "BLOQUE 1 — Datos del Curso")
    r += 1; field(r, "Nombre del módulo",    MODULO["nombre"])
    r += 1; field(r, "Código módulo",        MODULO["codigo"])
    r += 1; field(r, "Ciclo formativo",      MODULO["ciclo"])
    r += 1; field(r, "Curso",                MODULO["curso"])
    r += 1; field(r, "Año académico",        MODULO["anno"])
    r += 1; field(r, "Nombre del profesor/a","", editable=True)
    r += 1; field(r, "Grupo/s",              "", editable=True)
    r += 1; field(r, "Centro educativo",     "", editable=True)

    r += 2
    r = section(r, "⚙️", "BLOQUE 2 — Motor de Evaluación")
    r += 1; field(r, "N.º de evaluaciones",  MODULO["eval_count"], editable=False)
    r += 1; field(r, "Nota mínima aprobado", 5, editable=True)
    r += 1; field(r, "Nota máxima",          10, editable=False)
    r += 1; field(r, "Redondeo",             "0.00 decimales", editable=True)
    r += 1; field(r, "Peso Eval 1  (RA1 13% + RA2 19%)",          "32%", editable=False)
    r += 1; field(r, "Peso Eval 2  (RA3 9% + RA4 16% + RA5 13%)", "38%", editable=False)
    r += 1; field(r, "Peso Eval 3  (RA6 12% + RA7 13% + RA8 5%)", "30%", editable=False)
    # ── Pesos examen / prácticas (editables — se usan en Reg. Notas) ──
    r += 1
    c = ws.cell(r, 2, value="Peso examen (teoría)")
    c.font = fn(size=10, bold=True, color=NAVY); c.fill = f(ICE)
    c.alignment = al(v="center"); c.border = thin_border()
    val_ex = ws.cell(r, 3, value=PESO_EXAMEN)
    val_ex.number_format = "0%"
    val_ex.font = fn(size=10, bold=True, color="C00000")
    val_ex.fill = f("FDECEA"); val_ex.alignment = al(h="center",v="center")
    val_ex.border = thin_border()
    # Nombre de rango para referencias desde Reg. Notas
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.utils import quote_sheetname, absolute_coordinate
    cfg_sheet = "Configuración"
    wb.defined_names["PESO_EXAMEN"]   = DefinedName("PESO_EXAMEN",
        attr_text=f"{quote_sheetname(cfg_sheet)}!{absolute_coordinate(ws.cell(r,3).coordinate)}")
    r += 1
    c = ws.cell(r, 2, value="Peso prácticas (actividades)")
    c.font = fn(size=10, bold=True, color=NAVY); c.fill = f(ICE)
    c.alignment = al(v="center"); c.border = thin_border()
    val_pr = ws.cell(r, 3, value=PESO_PRACTICA)
    val_pr.number_format = "0%"
    val_pr.font = fn(size=10, bold=True, color=GREEN)
    val_pr.fill = f(GRLIT); val_pr.alignment = al(h="center",v="center")
    val_pr.border = thin_border()
    wb.defined_names["PESO_PRACTICA"] = DefinedName("PESO_PRACTICA",
        attr_text=f"{quote_sheetname(cfg_sheet)}!{absolute_coordinate(ws.cell(r,3).coordinate)}")
    nota_pesos = ws.cell(r+1, 2,
        value="⚠️  Los dos valores deben sumar 100%. Si los cambias, las notas de Reg. Notas se recalculan automáticamente.")
    nota_pesos.font = fn(size=8, italic=True, color="7F6000")
    nota_pesos.fill = f("FFF2CC")
    nota_pesos.alignment = al(h="left", v="center", wrap=True)
    ws.merge_cells(f"B{r+1}:C{r+1}")
    ws.row_dimensions[r+1].height = 28
    r += 1
    r += 1; field(r, "Recuperación",         "Sí — convocatoria extraordinaria", editable=True)
    r += 2
    # Nota informativa sobre prácticas en empresa
    nota_empresa = (
        "⚠️  PRÁCTICAS EN EMPRESA (2ª EVALUACIÓN): El alumnado realiza 18h de prácticas "
        "en empresa durante la 2ª evaluación (3 semanas × 6h/sem). Estas horas afectan "
        "únicamente a la temporalización — hay menos tiempo para impartir UT3, UT4 y UT5 — "
        "pero NO modifican las ponderaciones curriculares de los RAs."
    )
    c = ws.cell(r, 2, value=nota_empresa)
    c.font = fn(size=9, italic=True, color="7F6000")
    c.fill = f("FFF2CC")
    c.alignment = al(h="left", v="center", wrap=True)
    ws.merge_cells(f"B{r}:F{r}")
    ws.row_dimensions[r].height = 52

    r += 2
    r = section(r, "🖨️", "BLOQUE 3 — Preferencias del Sistema")
    r += 1; field(r, "Formato boletines",    "PDF · 1 por alumno/a", editable=True)
    r += 1; field(r, "Ruta exportación PDF", "", editable=True)
    r += 1; field(r, "Email destino",        "", editable=True)
    r += 1; field(r, "Logo del centro",      "(no configurado)", editable=True)

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: PROGRAMACIÓN (la clave — distribución RA/CE por UT)
# ═══════════════════════════════════════════════════════════════════════════════
def build_programacion(wb):
    # ── Columnas ──────────────────────────────────────────────────────────────
    # B=Eval, C=UT, D=Unidad de Trabajo, E=Horas, F=% Total,
    # G=RA, H=Criterios, I=N.ºCEs, J=% RA en UT, K=Horas RA,
    # L=% Pond. Global RA, M=Instrumento, N=Apuntes (hipervínculo)
    LAST_COL = 14   # columna N = 14
    LAST_LET = "N"

    ws = wb.create_sheet(_sn("Programación"))
    apply_standard_header(
        ws,
        f"📋 Programación Didáctica — {MODULO['abrev']} {MODULO['curso']}",
        f"Distribución de RAs y CEs por Unidad de Trabajo · {MODULO['nombre']} · {MODULO['anno']}",
        _sn("Programación"),
    )
    set_col_width(ws, {
        "A": "3",  "B": "6",  "C": "8",  "D": "35",
        "E": "7",  "F": "7",  "G": "8",  "H": "24",
        "I": "7",  "J": "9",  "K": "9",  "L": "10",
        "M": "22", "N": "16",
    })

    r = 8

    # ── Título sección ──────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"B{r}:{LAST_LET}{r}")
    tc = ws.cell(r, 2, value="MAPA DE UNIDADES DE TRABAJO · RESULTADOS DE APRENDIZAJE · CRITERIOS DE EVALUACIÓN")
    tc.font = fn(size=11, bold=True, color=WHITE)
    tc.fill = f(NAVY)
    tc.alignment = al(h="center", v="center")

    r += 1
    # Cabecera columnas
    ws.row_dimensions[r].height = 36
    headers = [
        (2,  "Eval."),
        (3,  "UT"),
        (4,  "Unidad de Trabajo"),
        (5,  "Horas"),
        (6,  "% Total"),
        (7,  "RA"),
        (8,  "Criterios de Evaluación (CEs)"),
        (9,  "N.º CEs"),
        (10, "% RA en UT"),
        (11, "Horas RA en UT"),
        (12, "% Pond. Global RA"),
        (13, "Instrumento(s)"),
        (14, "📖 Apuntes"),
    ]
    for col_i, label in headers:
        c = ws.cell(r, col_i, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(MED)
        c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # ── Filas de datos ───────────────────────────────────────────────────────
    for ut in UTS:
        ut_asigs = [(ra_id, ces) for u, ra_id, ces in ASIGNACIONES if u == ut["id"]]
        total_ces_ut = sum(len(ces) for _, ces in ut_asigs)
        ut_horas = ut["horas"]
        pond_ut  = ut_horas / TOTAL_HORAS * 100
        ut_eval  = ut.get("eval", "")
        ut_idx   = UTS.index(ut)
        bg_block = ICE if (ut_idx % 2 == 0) else WHITE
        eval_colors = {1: ("4472C4", "DDEEFF"), 2: ("E87222", "FDE9D7"), 3: ("70AD47", "E2EFDA")}
        ev_col, _ = eval_colors.get(ut_eval, (MED, ICE))

        for row_i, (ra_id, ces) in enumerate(ut_asigs):
            r += 1
            ra = RA_BY_ID[ra_id]
            ra_color, ra_lite = RA_COLORS[ra_id]
            is_first = (row_i == 0)
            bg = bg_block

            ws.row_dimensions[r].height = 22

            # Col B: Evaluación (solo primera fila del bloque UT)
            bc = ws.cell(r, 2, value=f"EV{ut_eval}" if is_first else "")
            bc.font = fn(size=9, bold=True, color=WHITE)
            bc.fill = f(ev_col if is_first else LGRAY)
            bc.alignment = al(h="center", v="center")
            bc.border = thin_border(WHITE)

            # Col C: UT id (solo primera fila del bloque)
            cc = ws.cell(r, 3, value=ut["id"] if is_first else "")
            cc.font = fn(size=10, bold=True, color=WHITE)
            cc.fill = f(NAVY)
            cc.alignment = al(h="center", v="center")
            cc.border = thin_border(WHITE)

            # Col D: Nombre UT (solo primera fila)
            dc = ws.cell(r, 4, value=ut["nombre"] if is_first else "")
            dc.font = fn(size=10, bold=is_first, color=NAVY if is_first else DGRAY)
            dc.fill = f(bg)
            dc.alignment = al(v="center", wrap=True)
            dc.border = thin_border()

            # Col E: Horas (solo primera fila)
            ec = ws.cell(r, 5, value=ut_horas if is_first else "")
            ec.font = fn(size=10, bold=is_first, color=DGRAY)
            ec.fill = f(bg)
            ec.alignment = al(h="center", v="center")
            ec.border = thin_border()

            # Col F: % Total (solo primera fila)
            fc_cell = ws.cell(r, 6, value=f"{pond_ut:.1f}%" if is_first else "")
            fc_cell.font = fn(size=10, color=DGRAY)
            fc_cell.fill = f(bg)
            fc_cell.alignment = al(h="center", v="center")
            fc_cell.border = thin_border()

            # Col G: RA id (badge coloreado; 🏢 si es dual)
            is_dual = (ra_id == DUAL_RA)
            ra_label = f"🏢 {ra_id}" if is_dual else ra_id
            gc = ws.cell(r, 7, value=ra_label)
            gc.font = fn(size=10, bold=True, color=WHITE)
            gc.fill = f("1F6B3A" if is_dual else ra_color)
            gc.alignment = al(h="center", v="center")
            gc.border = thin_border(WHITE)

            # Col H: CEs como texto
            hc = ws.cell(r, 8, value="  ".join(ces))
            hc.font = fn(size=9, color=DGRAY)
            hc.fill = f(ra_lite)
            hc.alignment = al(v="center", wrap=True)
            hc.border = thin_border()

            # Col I: N.º CEs
            ic = ws.cell(r, 9, value=len(ces))
            ic.font = fn(size=10, color=DGRAY)
            ic.fill = f(bg)
            ic.alignment = al(h="center", v="center")
            ic.border = thin_border()

            # Col J: % RA dentro del UT
            pct_ra_ut = len(ces) / total_ces_ut * 100
            jc = ws.cell(r, 10, value=f"{pct_ra_ut:.1f}%")
            jc.font = fn(size=10, color=DGRAY)
            jc.fill = f(bg)
            jc.alignment = al(h="center", v="center")
            jc.border = thin_border()

            # Col K: Horas RA en este UT
            h_ra_ut = ut_horas * len(ces) / total_ces_ut
            kc = ws.cell(r, 11, value=round(h_ra_ut, 1))
            kc.font = fn(size=10, color=DGRAY)
            kc.fill = f(bg)
            kc.alignment = al(h="center", v="center")
            kc.border = thin_border()

            # Col L: % Ponderación global del RA
            lc = ws.cell(r, 12, value=f"{ra['pond']}%")
            lc.font = fn(size=10, bold=True, color=NAVY)
            lc.fill = f(ra_lite)
            lc.alignment = al(h="center", v="center")
            lc.border = thin_border()

            # Col M: Instrumento(s) de evaluación (solo primera vez que aparece el RA en este UT)
            instrumentos = RA_INSTRUMENTOS.get(ra_id, [])
            label_map = {"examen": "📝 Examen", "practica": "🔧 Práctica", "empresa": "🏢 Empresa"}
            instr_str = "  ·  ".join(label_map.get(i, i) for i in instrumentos)
            mc = ws.cell(r, 13, value=instr_str)
            mc.font = fn(size=9, color=DGRAY, italic=True)
            mc.fill = f(ra_lite)
            mc.alignment = al(v="center", wrap=True)
            mc.border = thin_border()

            # Col N: Hipervínculo al apunte HTML de esta UT (solo primera fila del bloque UT)
            if is_first:
                apunte_url = _apunte_rel_path(ut)
                nc = ws.cell(r, 14, value=f'=HYPERLINK("{apunte_url}","📖 Ver apunte")')
                nc.font = fn(size=9, color="2563EB", italic=True)
                nc.fill = f("EBF5FE")
                nc.alignment = al(h="center", v="center")
                nc.border = thin_border()
            else:
                nc = ws.cell(r, 14, value="")
                nc.fill = f(bg)
                nc.border = thin_border()

        # Línea separadora entre UTs
        r += 1
        ws.row_dimensions[r].height = 4
        for ci in range(2, LAST_COL + 1):
            ws.cell(r, ci).fill = f(GOLD if (ut_idx % 2 == 0) else MGRAY)

    # ── Tabla resumen de RAs ─────────────────────────────────────────────────
    r += 2
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"B{r}:{LAST_LET}{r}")
    tc = ws.cell(r, 2, value="RESUMEN DE RESULTADOS DE APRENDIZAJE — PONDERACIÓN TOTAL DEL MÓDULO")
    tc.font = fn(size=11, bold=True, color=WHITE)
    tc.fill = f(NAVY)
    tc.alignment = al(h="center", v="center")

    r += 1
    ws.row_dimensions[r].height = 22
    sum_headers = [
        (2, "Eval."), (3, "RA"), (4, "Resultado de Aprendizaje"),
        (5, "Total CEs"), (6, "Horas equiv."), (7, "Ponderación %"), (8, "Instrumento(s)"),
    ]
    for col_i, h in sum_headers:
        c = ws.cell(r, col_i, value=h)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(MED)
        c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)
    # merge instrumento header across remaining cols
    ws.merge_cells(f"H{r}:{LAST_LET}{r}")

    # Determine which eval each RA belongs to (first eval that lists it)
    RA_EVAL = {}
    for ev, ra_ids in EVAL_RAS.items():
        for rid in ra_ids:
            RA_EVAL.setdefault(rid, ev)
    eval_colors = {1: ("4472C4", "DDEEFF"), 2: ("E87222", "FDE9D7"), 3: ("70AD47", "E2EFDA")}

    total_pond = 0
    for i, ra in enumerate(RAS):
        r += 1
        ws.row_dimensions[r].height = 20
        bg = ICE if i % 2 == 0 else WHITE
        ra_color, ra_lite = RA_COLORS[ra["id"]]
        n_ces = len(CES_POR_RA.get(ra["id"], []))
        h_ra  = HORAS_RA[ra["id"]]
        total_pond += ra["pond"]
        ev = RA_EVAL.get(ra["id"], "")
        ev_col, _ = eval_colors.get(ev, (MED, ICE))

        # Col B: Eval badge
        bc = ws.cell(r, 2, value=f"EV{ev}" if ev else "")
        bc.font = fn(size=9, bold=True, color=WHITE)
        bc.fill = f(ev_col)
        bc.alignment = al(h="center", v="center")
        bc.border = thin_border(WHITE)

        # Col C: RA badge
        is_dual = (ra["id"] == DUAL_RA)
        ra_label = f"🏢 {ra['id']}" if is_dual else ra["id"]
        rc = ws.cell(r, 3, value=ra_label)
        rc.font = fn(size=10, bold=True, color=WHITE)
        rc.fill = f("1F6B3A" if is_dual else ra_color)
        rc.alignment = al(h="center", v="center")
        rc.border = thin_border(WHITE)

        # Col D: nombre RA
        nc = ws.cell(r, 4, value=ra["nombre"])
        nc.font = fn(size=10, color=DGRAY)
        nc.fill = f(bg)
        nc.alignment = al(h="left", v="center", wrap=True)
        nc.border = thin_border()

        # Cols E-G: CEs, horas, ponderación
        for col_i, val in [(5, n_ces), (6, f"{h_ra:.1f} h"), (7, f'{ra["pond"]}%')]:
            c = ws.cell(r, col_i, value=val)
            c.font = fn(size=10, bold=(col_i == 7), color=NAVY if col_i == 7 else DGRAY)
            c.fill = f(ra_lite if col_i == 7 else bg)
            c.alignment = al(h="center", v="center")
            c.border = thin_border()

        # Col H-M: Instrumento(s) (merged)
        label_map = {"examen": "📝 Examen", "practica": "🔧 Práctica", "empresa": "🏢 Empresa"}
        instrumentos = RA_INSTRUMENTOS.get(ra["id"], [])
        instr_str = "  ·  ".join(label_map.get(i, i) for i in instrumentos)
        ws.merge_cells(f"H{r}:{LAST_LET}{r}")
        mc = ws.cell(r, 8, value=instr_str)
        mc.font = fn(size=9, color=DGRAY, italic=True)
        mc.fill = f(ra_lite)
        mc.alignment = al(v="center")
        mc.border = thin_border()

    # Fila total
    r += 1
    ws.row_dimensions[r].height = 20
    total_ces_all = sum(len(CES_POR_RA.get(ra["id"], [])) for ra in RAS)
    ws.merge_cells(f"B{r}:D{r}")
    c = ws.cell(r, 2, value=f"TOTAL — {len(RAS)} Resultados de Aprendizaje")
    c.font = fn(size=10, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(h="left", v="center")
    c.border = thin_border(WHITE)
    for col_i, val in [(5, total_ces_all), (6, f"{TOTAL_HORAS} h"), (7, f"{total_pond}%")]:
        c = ws.cell(r, col_i, value=val)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(NAVY); c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)
    ws.merge_cells(f"H{r}:{LAST_LET}{r}")
    ws.cell(r, 8).fill = f(NAVY)

    # ── Nota justificación DUAL ──────────────────────────────────────────────
    r += 2
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:{LAST_LET}{r}")
    c = ws.cell(r, 2, value="🏢  EVALUACIÓN DUAL — Justificación")
    c.font = fn(size=10, bold=True, color=WHITE)
    c.fill = f("1F6B3A"); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 80
    ws.merge_cells(f"B{r}:{LAST_LET}{r}")
    c = ws.cell(r, 2, value=DUAL_JUSTIFICACION)
    c.font = fn(size=9, italic=True, color=DGRAY)
    c.fill = f("E8F5E9")
    c.alignment = al(h="left", v="top", wrap=True)
    c.border = thin_border("1F6B3A")

    r += 1
    ws.row_dimensions[r].height = 16
    ws.merge_cells(f"B{r}:{LAST_LET}{r}")
    c = ws.cell(r, 2,
                value=f"Peso del RA dual en la nota final: {DUAL_PCT_NOTA*100:.0f}% "
                      f"= {DUAL_PCT_NOTA} puntos sobre 1  |  "
                      f"Evaluador: Tutor/a de empresa  |  "
                      f"Base legal: Ley Orgánica 3/2022, art. 42")
    c.font = fn(size=9, bold=True, color="1F6B3A")
    c.fill = f("E8F5E9"); c.alignment = al(v="center")

    ws.freeze_panes = "D10"
    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: ALUMNOS
# ═══════════════════════════════════════════════════════════════════════════════
def build_alumnos(wb):
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet(_sn("Alumnos"))
    apply_standard_header(ws, "👥 Lista de Alumnos",
                           "Registro de alumnado del grupo — datos académicos y de contacto",
                           _sn("Alumnos"))
    set_col_width(ws, {"A":"3","B":"5","C":"25","D":"20","E":"12",
                        "F":"15","G":"28","H":"28","I":"15","J":"12"})

    # Fila 8: contadores
    r = 8
    ws.row_dimensions[r].height = 18
    ws.cell(r, 2, value="Total activos:").font = fn(size=9, bold=True, color=NAVY)
    c = ws.cell(r, 3, value='=COUNTIF(J9:J38,"Activo")')
    c.font = fn(size=9, bold=True, color=GREEN)
    ws.cell(r, 4, value="Bajas:").font = fn(size=9, bold=True, color=NAVY)
    c = ws.cell(r, 5, value='=COUNTIF(J9:J38,"Baja")')
    c.font = fn(size=9, bold=True, color="C00000")
    ws.cell(r, 6, value="Pendientes:").font = fn(size=9, bold=True, color=NAVY)
    c = ws.cell(r, 7, value='=COUNTIF(J9:J38,"Pendiente")')
    c.font = fn(size=9, bold=True, color="7F6000")

    # Fila 9 (antes 8): cabeceras
    r = 9
    headers = ["Nº","Apellidos","Nombre","NIA / Exp.","Fecha nacim.",
               "Email","Teléfono contacto","Observaciones","Estado"]
    for ci, h in enumerate(headers):
        c = ws.cell(r, 2+ci, value=h)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(NAVY)
        c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)
    ws.row_dimensions[r].height = 22
    ws.freeze_panes = "B10"

    # Validación de datos para columna Estado (J = col 10)
    dv_estado = DataValidation(
        type="list",
        formula1='"Activo,Pendiente,Baja"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_estado.error      = "Usa el desplegable: Activo, Pendiente o Baja"
    dv_estado.errorTitle = "Valor no válido"
    dv_estado.prompt     = "Selecciona el estado del alumno/a"
    ws.add_data_validation(dv_estado)
    dv_estado.sqref = "J10:J39"

    for i in range(1, 31):
        r += 1   # filas 10-39
        bg = ICE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 18
        # Nº
        ws.cell(r, 2, value=i).fill = f(LGRAY)
        ws.cell(r, 2).font = fn(size=10, color=DGRAY, bold=True)
        ws.cell(r, 2).alignment = al(h="center", v="center")
        ws.cell(r, 2).border = thin_border()
        # Campos de datos (cols 3-9)
        for ci in range(3, 10):
            c = ws.cell(r, ci)
            c.fill = f(bg); c.border = thin_border(); c.alignment = al(v="center")
        # Estado (col 10) — dropdown
        c = ws.cell(r, 10, value="Activo")
        c.font = fn(size=9, color=GREEN)
        c.fill = f(GRLIT); c.border = thin_border()
        c.alignment = al(h="center", v="center")
    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: ACTIVIDADES
# ═══════════════════════════════════════════════════════════════════════════════
def build_actividades(wb):
    # Columnas:
    # B=ID_Act · C=Eval · D=UT · E=RA · F=CEs vinculados · G=Descripción
    # H=Instrumento · I=Tipo · J=Peso(%) · K=Nota máx.
    EVAL_COLORS = {1: "4472C4", 2: "E87222", 3: "70AD47"}
    EXAM_COLOR  = "C00000"

    ws = wb.create_sheet(_sn("Actividades"))
    apply_standard_header(
        ws,
        f"📝 Actividades Evaluables — {MODULO['abrev']} {MODULO['curso']}",
        f"Prácticas y exámenes vinculados a RAs y CEs · {MODULO['nombre']} · {MODULO['anno']}",
        _sn("Actividades"),
    )
    set_col_width(ws, {
        "A": "3",  "B": "8",  "C": "6",  "D": "7",
        "E": "7",  "F": "26", "G": "38",
        "H": "14", "I": "12", "J": "8",  "K": "9",
    })

    r = 8
    ws.row_dimensions[r].height = 14
    ws.cell(r, 2, value=(
        "ℹ️  Las actividades prácticas se han generado automáticamente desde la Programación. "
        "Ajusta la descripción, el tipo y el peso según tu planificación."
    )).font = fn(size=9, italic=True, color=DGRAY)

    r += 1
    ws.row_dimensions[r].height = 34
    hdr_cols = [
        (2, "ID"),
        (3, "Eval."),
        (4, "UT"),
        (5, "RA"),
        (6, "CEs vinculados"),
        (7, "Descripción de la actividad"),
        (8, "Instrumento"),
        (9, "Tipo"),
        (10, "Peso (%)"),
        (11, "Nota máx."),
    ]
    for col_i, label in hdr_cols:
        c = ws.cell(r, col_i, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(NAVY)
        c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)
    ws.freeze_panes = f"B{r+1}"

    # ── Bloque prácticas (generado desde ACTIVIDADES) ────────────────────────
    for i, act in enumerate(ACTIVIDADES):
        r += 1
        ev = act["eval"]
        ra_color, ra_lite = RA_COLORS[act["ra"]]
        bg = ra_lite
        ev_col = EVAL_COLORS.get(ev, MED)
        ws.row_dimensions[r].height = 20

        ces_str = "  ".join(act["ces"][:8])   # máx 8 CEs en la celda; el resto caben en texto largo
        if len(act["ces"]) > 8:
            ces_str += f"  … (+{len(act['ces'])-8})"

        row_vals = [
            (2,  act["id"],       NAVY,    "center", True,  f(LGRAY)),
            (3,  f"EV{ev}",       WHITE,   "center", True,  f(ev_col)),
            (4,  act["ut"],       WHITE,   "center", True,  f(NAVY)),
            (5,  act["ra"],       WHITE,   "center", True,  f(ra_color)),
            (6,  ces_str,         DGRAY,   "left",   False, f(bg)),
            (7,  act["desc"],     DGRAY,   "left",   False, f(WHITE)),
            (8,  "🔧 Práctica",   DGRAY,   "center", False, f(ICE)),
            (9,  "Individual",    DGRAY,   "center", False, f(ICE)),
            (10, "",              DGRAY,   "center", False, f(WHITE)),
            (11, 10,              NAVY,    "center", False, f(ICE)),
        ]
        for col_i, val, fc, ha, bld, fill in row_vals:
            c = ws.cell(r, col_i, value=val)
            c.font = fn(size=9, bold=bld, color=fc)
            c.fill = fill
            c.alignment = al(h=ha, v="center", wrap=(col_i in (6, 7)))
            c.border = thin_border()

    # ── Separador ────────────────────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 6
    for ci in range(2, 12):
        ws.cell(r, ci).fill = f(GOLD)

    # ── Bloque exámenes (uno por evaluación) ─────────────────────────────────
    for ev in sorted(EVAL_RAS.keys()):
        r += 1
        ws.row_dimensions[r].height = 20
        ra_ids_ev = EVAL_RAS[ev]
        uts_ev    = [ut["id"] for ut in UTS if ut.get("eval") == ev]
        ev_col    = EVAL_COLORS.get(ev, MED)

        row_vals = [
            (2,  f"EX{ev}",                       WHITE,   "center", True,  f(EXAM_COLOR)),
            (3,  f"EV{ev}",                        WHITE,   "center", True,  f(ev_col)),
            (4,  "+".join(uts_ev),                 WHITE,   "center", False, f(NAVY)),
            (5,  "+".join(ra_ids_ev),              WHITE,   "center", False, f(EXAM_COLOR)),
            (6,  "Todos los CEs de la evaluación", DGRAY,   "left",   False, f(ICE)),
            (7,  f"Examen teórico-práctico Evaluación {ev}", DGRAY, "left", False, f(WHITE)),
            (8,  "📝 Examen",                      DGRAY,   "center", False, f(ICE)),
            (9,  "Individual",                     DGRAY,   "center", False, f(ICE)),
            (10, "",                               DGRAY,   "center", False, f(WHITE)),
            (11, 10,                               NAVY,    "center", False, f(ICE)),
        ]
        for col_i, val, fc, ha, bld, fill in row_vals:
            c = ws.cell(r, col_i, value=val)
            c.font = fn(size=9, bold=bld, color=fc)
            c.fill = fill
            c.alignment = al(h=ha, v="center", wrap=(col_i == 7))
            c.border = thin_border()

    # ── Nota al pie ──────────────────────────────────────────────────────────
    r += 2
    ws.merge_cells(f"B{r}:K{r}")
    c = ws.cell(r, 2, value=(
        "💡  Peso (%): indica el porcentaje de la nota de prácticas que pondera esta actividad "
        "para cada RA. Los exámenes tienen su propio peso (configurable en Configuración → Peso examen). "
        "La suma de pesos de prácticas por RA debería ser 100%."
    ))
    c.font = fn(size=8, italic=True, color=DGRAY)
    c.fill = f(LGRAY)
    c.alignment = al(h="left", v="center", wrap=True)
    ws.row_dimensions[r].height = 28

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: REG. NOTAS
# ═══════════════════════════════════════════════════════════════════════════════
# Pesos teoría / práctica (del cuaderno histórico del profesor)
PESO_EXAMEN   = 0.75
PESO_PRACTICA = 0.25

# Examen que corresponde a cada evaluación
# EX1 → Eval 1 (RA1, RA2)  |  EX2 → Eval 2 (RA3,RA4,RA5)  |  EX3 → Eval 3 (RA6,RA7,RA8)
RA_EXAMEN = {ra: f"EX{ev}" for ev, ras in EVAL_RAS.items() for ra in ras}

def build_reg_notas(wb):
    ws = wb.create_sheet(_sn("Reg. Notas"))
    apply_standard_header(
        ws,
        "📊 Registro de Notas",
        f"Calificaciones por alumno/a · Actividades, Exámenes y RAs · "
        f"{MODULO['abrev']} {MODULO['curso']} · {MODULO['anno']}",
        _sn("Reg. Notas"),
    )

    r = 8
    ws.cell(r, 2, value=(
        "ℹ️  Introduzca las calificaciones de actividades (prácticas) y exámenes. "
        "Nota RA = % examen + % prácticas (configurables en la hoja Configuración). "
        "Las columnas de RAs y la nota final se calculan automáticamente."
    )).font = fn(size=9, italic=True, color=DGRAY)
    ws.row_dimensions[r].height = 16

    # ── Layout de columnas (calculado dinámicamente desde ACTIVIDADES) ───────
    # Col 2  (B): Nº
    # Col 3  (C): Nombre
    # Col 4..    : Actividades A001-A00N  (len(ACTIVIDADES) cols)
    # Col ..     : Exámenes EX1..EXM      (eval_count cols)
    # Col ..     : RA1-RAN               (len(RAS) cols)
    # Col ..     : NOTA FINAL
    ACT_START = 4
    act_ids   = [a["id"] for a in ACTIVIDADES]
    EX_START  = ACT_START + len(act_ids)
    ex_ids    = [f"EX{ev}" for ev in sorted(EVAL_RAS.keys())]
    RA_START  = EX_START + len(ex_ids)
    ra_ids    = [ra["id"] for ra in RAS]
    NF_COL    = RA_START + len(RAS)

    # Mapa dinámico: ra_id → lista de columnas de actividades prácticas
    ra_act_map: dict[str, list[int]] = {}
    for i, act in enumerate(ACTIVIDADES):
        ra_act_map.setdefault(act["ra"], []).append(ACT_START + i)

    # Mapa examen: "EX1" → columna absoluta
    ex_col_map = {f"EX{ev}": EX_START + i for i, ev in enumerate(sorted(EVAL_RAS.keys()))}

    # Fila de sub-cabeceras de grupo (fila 9)
    r = 9
    ws.row_dimensions[r].height = 14
    grupos = [
        (ACT_START, len(act_ids), "ACTIVIDADES (PRÁCTICAS)", NAVY),
        (EX_START,  len(ex_ids),  "EXÁMENES (TEORÍA)",       "C00000"),
        (RA_START,  len(ra_ids),  "NOTA RA  (75% EX + 25% ACT)", PURP),
        (NF_COL,    1,            "FINAL",                   ORNG),
    ]
    for col_ini, n_cols, label, color in grupos:
        ws.merge_cells(start_row=r, start_column=col_ini,
                       end_row=r,   end_column=col_ini+n_cols-1)
        c = ws.cell(r, col_ini, value=label)
        c.font = fn(size=8, bold=True, color=WHITE)
        c.fill = f(color)
        c.alignment = al(h="center", v="center")

    # Fila de cabeceras de columna (fila 10)
    r = 10
    ws.row_dimensions[r].height = 18
    headers = ["Nº","Apellidos y Nombre"] + act_ids + ex_ids + ra_ids + ["NOTA FINAL"]
    for ci, h in enumerate(headers):
        col = 2 + ci
        c = ws.cell(r, col, value=h)
        if h in ra_ids:
            bg = PURP
        elif h in ex_ids:
            bg = "C00000"
        elif h == "NOTA FINAL":
            bg = ORNG
        else:
            bg = NAVY
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg)
        c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    ws.freeze_panes = "D11"

    # Filas de datos (alumnos)
    for i in range(1, 31):
        r += 1
        bg = ICE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 18

        # Nº
        c = ws.cell(r, 2, value=i)
        c.font = fn(size=9, bold=True, color=DGRAY)
        c.fill = f(LGRAY); c.alignment = al(h="center",v="center"); c.border = thin_border()

        # Nombre — auto desde Alumnos
        c = ws.cell(r, 3, value=nombre_formula(r))
        c.fill = f(ICE); c.border = thin_border(); c.alignment = al(v="center")
        c.font = fn(size=9, color=NAVY)

        # Actividades — entrada manual (blanco)
        for ai in range(len(act_ids)):
            c = ws.cell(r, ACT_START+ai)
            c.fill = f(WHITE); c.border = thin_border()
            c.alignment = al(h="center",v="center")

        # Exámenes — entrada manual (rojo muy claro)
        for ei in range(len(ex_ids)):
            c = ws.cell(r, EX_START+ei)
            c.fill = f("FDECEA"); c.border = thin_border()
            c.alignment = al(h="center",v="center")

        # Nota RA = 75% examen_eval + 25% promedio_actividades
        for ri, ra in enumerate(RAS):
            col = RA_START + ri
            act_cols = ra_act_map.get(ra["id"], [])
            ex_col   = ex_col_map[RA_EXAMEN[ra["id"]]]
            ex_ref   = f"{get_column_letter(ex_col)}{r}"
            act_refs = ",".join([f"{get_column_letter(c)}{r}" for c in act_cols])
            formula  = (
                f'=IFERROR(ROUND(PESO_EXAMEN*{ex_ref}'
                f'+PESO_PRACTICA*AVERAGE({act_refs}),2),"")'
            )
            c = ws.cell(r, col, value=formula)
            _, ra_lite = RA_COLORS[ra["id"]]
            c.fill = f(ra_lite); c.border = thin_border()
            c.alignment = al(h="center",v="center")
            c.font = fn(size=9, color=NAVY)

        # NOTA FINAL — suma ponderada de RAs
        ra_col_weights = [
            (get_column_letter(RA_START + ri), RAS[ri]["pond"] / 100)
            for ri in range(len(RAS))
        ]
        c = ws.cell(r, NF_COL, value=weighted_formula(ra_col_weights, r))
        c.fill = f(ICE); c.border = thin_border()
        c.alignment = al(h="center",v="center")
        c.font = fn(size=9, bold=True, color=NAVY)

    # Anchos de columna
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 22
    for ci in range(len(act_ids)):
        ws.column_dimensions[get_column_letter(ACT_START+ci)].width = 8
    for ci in range(len(ex_ids)):
        ws.column_dimensions[get_column_letter(EX_START+ci)].width = 9
    for ci in range(len(ra_ids)):
        ws.column_dimensions[get_column_letter(RA_START+ci)].width = 8
    ws.column_dimensions[get_column_letter(NF_COL)].width = 10

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: RESUMEN
# ═══════════════════════════════════════════════════════════════════════════════
def build_resumen(wb):
    ws = wb.create_sheet(_sn("Resumen"))
    apply_standard_header(
        ws,
        "📈 Resumen y Dashboard",
        f"Vista general del grupo · Estadísticas y KPIs · {MODULO['abrev']} {MODULO['curso']} · {MODULO['anno']}",
        _sn("Resumen"),
    )
    set_col_width(ws, {"A":"3","B":"20","C":"15","D":"15","E":"15",
                        "F":"15","G":"15","H":"15","I":"15"})
    r = 8
    ws.cell(r, 2, value="🚧  Este panel se completará con datos de Reg. Notas en Sprint 9 (Dashboard).").font = fn(size=10, italic=True, color=DGRAY)

    r += 2
    kpis = [
        ("📊 Alumnos registrados", f"=COUNTA({_sr('Alumnos')}!C9:C38)"),
        ("✅ Aprobados (≥5)", "—"),
        ("❌ Suspensos (<5)", "—"),
        ("📈 Media del grupo",  "—"),
        ("⚠️ En riesgo (<4)",   "—"),
    ]
    for i, (label, val) in enumerate(kpis):
        c = ws.cell(r+i, 2, value=label)
        c.font = fn(size=11, bold=True, color=NAVY)
        c.fill = f(ICE); c.alignment = al(v="center"); c.border = thin_border()
        v = ws.cell(r+i, 3, value=val)
        v.font = fn(size=12, bold=True, color=NAVY)
        v.fill = f(WHITE); v.alignment = al(h="center",v="center"); v.border = thin_border()
        ws.row_dimensions[r+i].height = 24
    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: EVALUACIÓN (1, 2, 3)
# ═══════════════════════════════════════════════════════════════════════════════
def build_evaluacion(wb, num):
    name = f"Evaluación {num}"
    ws = wb.create_sheet(_sn(name))
    ev_color = {1:"4472C4", 2:"70AD47", 3:"E87222"}[num]
    uts_eval  = [ut for ut in UTS if ut["eval"] == num]
    ra_ids_ev = EVAL_RAS[num]                        # Solo RAs de esta eval
    nw        = pond_norm_eval(num)                  # Pesos normalizados {ra_id: w}

    # Subtítulo con ponderaciones normalizadas
    pond_str = "  ·  ".join(f'{rid} {nw[rid]*100:.1f}%' for rid in ra_ids_ev)
    apply_standard_header(
        ws,
        f"📝 Evaluación {num}  ·  {', '.join(u['id'] for u in uts_eval)}",
        f"RAs evaluados: {pond_str}  (pesos normalizados al 100% de esta evaluación)",
        _sn(name))

    # Columnas: B=Nº, C=Nombre, D..=UT cols, luego RA cols, luego NOTA EVAL, RESULTADO
    UT_START = 4
    RA_START = UT_START + len(uts_eval)
    FINAL_COL = RA_START + len(ra_ids_ev)
    RES_COL   = FINAL_COL + 1

    widths = {"A":"3","B":"5","C":"30"}
    for i in range(len(uts_eval)):
        widths[get_column_letter(UT_START + i)] = "10"
    for i in range(len(ra_ids_ev)):
        widths[get_column_letter(RA_START + i)] = "10"
    widths[get_column_letter(FINAL_COL)] = "12"
    widths[get_column_letter(RES_COL)]   = "12"
    set_col_width(ws, widths)

    # Cabecera fila 8: leyenda pesos
    r = 8
    ws.row_dimensions[r].height = 14
    for i, rid in enumerate(ra_ids_ev):
        c = ws.cell(r, RA_START + i,
                    value=f'{rid}: {RA_BY_ID[rid]["pond"]}% global → {nw[rid]*100:.1f}% en Eval {num}')
        ra_color, _ = RA_COLORS[rid]
        c.font = fn(size=8, italic=True, color=ra_color)

    # Cabecera tabla fila 10
    r = 10
    ws.row_dimensions[r].height = 36

    for ci, label in [(2,"Nº"),(3,"Apellidos y Nombre")]:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(NAVY); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # UT columns
    for i, ut in enumerate(uts_eval):
        c = ws.cell(r, UT_START + i, value=f'{ut["id"]}\n({ut["horas"]}h)')
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(ev_color); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # RA columns (solo los de esta eval) con ponderación normalizada; dual en verde empresa
    for i, rid in enumerate(ra_ids_ev):
        ra_color, _ = RA_COLORS[rid]
        is_dual = (rid == DUAL_RA)
        label = (f'🏢 {rid}\n({nw[rid]*100:.1f}%)\nEMPRESA'
                 if is_dual else f'{rid}\n({nw[rid]*100:.1f}%)')
        bg_col = "1F6B3A" if is_dual else ra_color
        c = ws.cell(r, RA_START + i, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg_col); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # NOTA EVAL y RESULTADO
    for ci, label, bg in [(FINAL_COL,"NOTA EVAL",NAVY),(RES_COL,"Resultado",NAVY)]:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    ws.freeze_panes = f"D{r+1}"

    # Filas de alumnos
    for i in range(1, 31):
        r += 1
        bg = ICE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20

        # Nº
        c = ws.cell(r, 2, value=i)
        c.font = fn(size=9, bold=True, color=DGRAY)
        c.fill = f(LGRAY); c.alignment = al(h="center",v="center"); c.border = thin_border()

        # Nombre — auto desde Alumnos
        c = ws.cell(r, 3, value=nombre_formula(r))
        c.fill = f(ICE); c.border = thin_border(); c.alignment = al(v="center")
        c.font = fn(size=9, color=NAVY)

        # UT grades (entrada manual)
        for j in range(len(uts_eval)):
            c = ws.cell(r, UT_START + j)
            c.fill = f(WHITE); c.border = thin_border()
            c.alignment = al(h="center", v="center")
            c.number_format = "0.00"

        # RA grades — fórmula que tira de Reg. Notas (hereda PESO_EXAMEN/PESO_PRACTICA)
        for j, rid in enumerate(ra_ids_ev):
            _, ra_lite = RA_COLORS[rid]
            is_dual = (rid == DUAL_RA)
            cell_bg = "C8E6C9" if is_dual else ra_lite
            rn_col  = RN_RA_COL[rid]
            c = ws.cell(r, RA_START + j,
                        value=f"=IFERROR({_sr('Reg. Notas')}!{rn_col}{r},\"\")")
            c.fill = f(cell_bg); c.border = thin_border()
            c.alignment = al(h="center", v="center")
            c.number_format = "0.00"
            c.font = fn(size=9, color="1F6B3A" if is_dual else NAVY)

        # NOTA EVAL = suma ponderada (pesos normalizados) de los RAs de esta eval
        ra_col_weights = [
            (get_column_letter(RA_START + j), nw[rid])
            for j, rid in enumerate(ra_ids_ev)
        ]
        c = ws.cell(r, FINAL_COL, value=weighted_formula(ra_col_weights, r))
        c.font = fn(size=9, bold=True, color=NAVY)
        c.fill = f(ICE); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.number_format = "0.00"

        # Resultado
        nf_ref = f"{get_column_letter(FINAL_COL)}{r}"
        c = ws.cell(r, RES_COL,
                    value=f'=IF({nf_ref}="","",IF({nf_ref}>=5,"APTO","NO APTO"))')
        c.font = fn(size=9, bold=True, color=GREEN)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")

    return ws

# ═══════════════════════════════════════════════════════════════════════════════
# HOJAS OCULTAS DE DATOS
# ═══════════════════════════════════════════════════════════════════════════════
def build_data_sheets(wb):
    # ── _UTs ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet(_sn("_UTs"))
    ws.sheet_state = "hidden"
    headers = ["id","nombre","horas","evaluacion","pond_pct","tags"]
    for ci, h in enumerate(headers):
        ws.cell(1, ci+1, value=h).font = fn(bold=True, color=WHITE)
        ws.cell(1, ci+1).fill = f(NAVY)
    for ri, ut in enumerate(UTS):
        pond = ut["horas"] / TOTAL_HORAS * 100
        row_data = [ut["id"], ut["nombre"], ut["horas"], ut["eval"], round(pond,2), ut["tags"]]
        for ci, val in enumerate(row_data):
            ws.cell(ri+2, ci+1, value=val)

    # ── _RAs ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet(_sn("_RAs"))
    ws.sheet_state = "hidden"
    headers = ["id","nombre","pond_pct","total_ces","horas_equiv"]
    for ci, h in enumerate(headers):
        ws.cell(1, ci+1, value=h).font = fn(bold=True, color=WHITE)
        ws.cell(1, ci+1).fill = f(NAVY)
    for ri, ra in enumerate(RAS):
        n_ces = len(CES_POR_RA.get(ra["id"], []))
        h_ra  = HORAS_RA[ra["id"]]
        row_data = [ra["id"], ra["nombre"], ra["pond"], n_ces, round(h_ra,2)]
        for ci, val in enumerate(row_data):
            ws.cell(ri+2, ci+1, value=val)

    # ── _CEs ──────────────────────────────────────────────────────────────────
    ws = wb.create_sheet(_sn("_CEs"))
    ws.sheet_state = "hidden"
    headers = ["id_ra","id_ce","ce_numero"]
    for ci, h in enumerate(headers):
        ws.cell(1, ci+1, value=h).font = fn(bold=True, color=WHITE)
        ws.cell(1, ci+1).fill = f(NAVY)
    row = 2
    for ra in RAS:
        for ce in CES_POR_RA.get(ra["id"], []):
            ws.cell(row, 1, value=ra["id"])
            ws.cell(row, 2, value=ce)
            ws.cell(row, 3, value=int(ce[2:]))
            row += 1

    # ── _Asignaciones ─────────────────────────────────────────────────────────
    ws = wb.create_sheet(_sn("_Asignaciones"))
    ws.sheet_state = "hidden"
    headers = ["id_ut","id_ra","ces","n_ces","horas_ut","pond_ut_pct","h_ra_en_ut"]
    for ci, h in enumerate(headers):
        ws.cell(1, ci+1, value=h).font = fn(bold=True, color=WHITE)
        ws.cell(1, ci+1).fill = f(NAVY)
    row = 2
    for ut_id, ra_id, ces in ASIGNACIONES:
        total_ces_ut = sum(len(c) for u, r, c in ASIGNACIONES if u == ut_id)
        h_ut  = UT_BY_ID[ut_id]["horas"]
        h_ra  = h_ut * len(ces) / total_ces_ut
        pond  = h_ut / TOTAL_HORAS * 100
        ws.cell(row, 1, value=ut_id)
        ws.cell(row, 2, value=ra_id)
        ws.cell(row, 3, value=", ".join(ces))
        ws.cell(row, 4, value=len(ces))
        ws.cell(row, 5, value=h_ut)
        ws.cell(row, 6, value=round(pond, 2))
        ws.cell(row, 7, value=round(h_ra, 2))
        row += 1

    # ── _Actividades — lista de actividades calculadas desde ASIGNACIONES (Sprint 7) ─
    ws = wb.create_sheet(_sn("_Actividades"))
    ws.sheet_state = "hidden"
    headers_act = ["id_actividad","id_ut","id_ra","evaluacion","n_ces","ces","descripcion","tipo"]
    for ci, h in enumerate(headers_act):
        c = ws.cell(1, ci+1, value=h)
        c.font = fn(bold=True, color=WHITE)
        c.fill = f(NAVY)
    for ri, act in enumerate(ACTIVIDADES):
        row_data = [
            act["id"], act["ut"], act["ra"], act["eval"],
            len(act["ces"]), ", ".join(act["ces"]),
            act["desc"], "Práctica",
        ]
        for ci, val in enumerate(row_data):
            ws.cell(ri+2, ci+1, value=val)
    # Exámenes
    for ev in sorted(EVAL_RAS.keys()):
        ri += 1
        uts_ev = [ut["id"] for ut in UTS if ut.get("eval") == ev]
        ra_ev  = EVAL_RAS[ev]
        ws.cell(ri+2, 1, value=f"EX{ev}")
        ws.cell(ri+2, 4, value=ev)
        ws.cell(ri+2, 5, value=0)
        ws.cell(ri+2, 7, value=f"Examen Evaluación {ev}")
        ws.cell(ri+2, 8, value="Examen")

    # ── _Alumnos — espejo relacional de la hoja Alumnos (Sprint 2) ────────────
    # Alumnos (Sprint 3): cabecera fila 9, datos filas 10-39 → alumno i en fila 9+i
    ws = wb.create_sheet(_sn("_Alumnos"))
    ws.sheet_state = "hidden"
    headers_al = ["id_alumno","apellidos","nombre","nia_expediente",
                  "fecha_nacimiento","email","telefono","observaciones","estado"]
    for ci, h in enumerate(headers_al):
        ws.cell(1, ci+1, value=h).font = fn(bold=True, color=WHITE)
        ws.cell(1, ci+1).fill = f(NAVY)
    for i in range(1, 31):
        al_row  = 9 + i   # fila en hoja Alumnos (Sprint 4: corregido de 8+i a 9+i)
        data_row = i + 1  # fila en _Alumnos (fila 1 = cabecera)
        ws.cell(data_row, 1, value=i)  # id_alumno = número de orden
        for ci, src_col in enumerate(range(3, 11)):   # cols C→J de Alumnos
            ws.cell(data_row, ci+2,
                    value=f"=IFERROR({_sr('Alumnos')}!{get_column_letter(src_col)}{al_row},\"\")")

    # ── _Grupos — datos del grupo/clase (Sprint 2) ────────────────────────────
    ws = wb.create_sheet(_sn("_Grupos"))
    ws.sheet_state = "hidden"
    headers_gr = ["id_grupo","nombre_grupo","modulo_codigo","modulo_nombre",
                  "curso_academico","horas_semanales","horas_totales","profesor","aula"]
    for ci, h in enumerate(headers_gr):
        ws.cell(1, ci+1, value=h).font = fn(bold=True, color=WHITE)
        ws.cell(1, ci+1).fill = f(NAVY)
    grupo_data = [
        "GR01",
        f"{MODULO['curso']} — {MODULO['abrev']}",
        MODULO["codigo"],
        MODULO["nombre"],
        MODULO["anno"],
        MODULO["horas_sem"],
        MODULO["total_horas"],
        "=IFERROR(Configuración!C11,\"\")",  # nombre profesor desde Configuración
        "",
    ]
    for ci, val in enumerate(grupo_data):
        ws.cell(2, ci+1, value=val)

    # ── _Notas_Actividades — tabla de notas en formato largo (Sprint 2) ───────
    # Esquema relacional: una fila por (alumno, actividad). Se rellena con macros
    # o manualmente. Sirve de base para el Dashboard (Sprint 9) y Auditoría (Sprint 11).
    ws = wb.create_sheet(_sn("_Notas_Actividades"))
    ws.sheet_state = "hidden"
    headers_na = ["id_alumno","id_actividad","id_ut","id_ra","nota",
                  "fecha_calificacion","convocatoria","observaciones"]
    for ci, h in enumerate(headers_na):
        c = ws.cell(1, ci+1, value=h)
        c.font = fn(bold=True, color=WHITE)
        c.fill = f(NAVY)
    ws.cell(2, 1, value="← Esta tabla se poblará automáticamente en Sprint 9 (Dashboard)")

    # ── _Evaluaciones — resultados por alumno y evaluación (Sprint 6) ─────────
    # Una fila por (alumno, evaluación): 30 alumnos × 3 evals = 90 filas de datos
    ws = wb.create_sheet(_sn("_Evaluaciones"))
    ws.sheet_state = "hidden"
    headers_ev = ["id_alumno","evaluacion","nota_eval","resultado","fecha_cierre","observaciones"]
    for ci, h in enumerate(headers_ev):
        c = ws.cell(1, ci+1, value=h)
        c.font = fn(bold=True, color=WHITE)
        c.fill = f(NAVY)

    # Mapa: evaluacion → columna de NOTA EVAL en la hoja Evaluación N
    # build_evaluacion layout: UT_START=4, RA_START=4+len(uts_eval), FINAL_COL=RA_START+len(ra_ids_ev)
    # Filas de alumnos en Evaluación N: cabecera fila 10 → datos filas 11-40
    def _ev_nota_col(num):
        uts_ev = [ut for ut in UTS if ut["eval"] == num]
        ra_ids_ev = EVAL_RAS[num]
        final_col = 4 + len(uts_ev) + len(ra_ids_ev)
        return get_column_letter(final_col)

    def _ev_res_col(num):
        uts_ev = [ut for ut in UTS if ut["eval"] == num]
        ra_ids_ev = EVAL_RAS[num]
        final_col = 4 + len(uts_ev) + len(ra_ids_ev)
        return get_column_letter(final_col + 1)

    data_row = 2
    for i in range(1, 31):    # alumnos 1-30
        for ev in sorted(EVAL_RAS.keys()):
            ev_nota_col = _ev_nota_col(ev)
            ev_res_col  = _ev_res_col(ev)
            al_row_ev   = 10 + i   # alumno i en fila 10+i de la hoja Evaluación N
            ws.cell(data_row, 1, value=i)
            ws.cell(data_row, 2, value=ev)
            ws.cell(data_row, 3,
                    value=f"=IFERROR({_sr(f'Evaluación {ev}')}!{ev_nota_col}{al_row_ev},\"\")")
            ws.cell(data_row, 4,
                    value=f"=IFERROR({_sr(f'Evaluación {ev}')}!{ev_res_col}{al_row_ev},\"\")")
            ws.cell(data_row, 5, value="")   # fecha_cierre: entrada manual
            ws.cell(data_row, 6, value="")   # observaciones
            data_row += 1

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: BOLETÍN INDIVIDUAL — Sprint 3
# ═══════════════════════════════════════════════════════════════════════════════
def build_boletin(wb):
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet(_sn("Boletín"))
    apply_standard_header(
        ws,
        "📄 Boletín Individual",
        f"Informe de calificaciones por alumno/a · {MODULO['abrev']} {MODULO['curso']} · {MODULO['anno']}",
        _sn("Boletín"),
    )
    set_col_width(ws, {
        "A":"3","B":"18","C":"14","D":"14","E":"14",
        "F":"14","G":"14","H":"14","I":"14","J":"14","K":"14","L":"14",
    })

    EV_COLORS  = {1: "4472C4", 2: "70AD47", 3: "E87222"}
    EV_EMOJIS  = {1: "📘", 2: "📗", 3: "📙"}
    LAST_COL_L = "L"   # columna L = 12, suficiente para cualquier layout

    # ── Selector de alumno/a ─────────────────────────────────────────────────
    r = 8
    ws.row_dimensions[r].height = 26
    c = ws.cell(r, 2, value="👤  Seleccionar alumno/a (número):")
    c.font = fn(size=11, bold=True, color=NAVY)
    c.fill = f(ICE); c.alignment = al(v="center"); c.border = thin_border()
    sel = ws.cell(r, 3, value=1)
    sel.font = fn(size=14, bold=True, color="C00000")
    sel.fill = f("FDECEA"); sel.alignment = al(h="center", v="center")
    sel.border = thin_border("C00000")
    ws.merge_cells("D8:H8")
    nom = ws.cell(r, 4,
        value=f"=IFERROR(INDEX({_sr('Alumnos')}!C:C,9+C8)&\", \"&INDEX({_sr('Alumnos')}!D:D,9+C8),\"\")")
    nom.font = fn(size=13, bold=True, color=NAVY)
    nom.fill = f(ICE); nom.alignment = al(v="center"); nom.border = thin_border()

    # Datos personales
    r = 9
    ws.row_dimensions[r].height = 18
    for ci, (label, formula) in enumerate([
        ("NIA/Exp.",   f"=IFERROR(INDEX({_sr('Alumnos')}!E:E,9+C8),\"\")"),
        ("Email",      f"=IFERROR(INDEX({_sr('Alumnos')}!G:G,9+C8),\"\")"),
        ("Teléfono",   f"=IFERROR(INDEX({_sr('Alumnos')}!H:H,9+C8),\"\")"),
        ("Estado",     f"=IFERROR(INDEX({_sr('Alumnos')}!J:J,9+C8),\"\")"),
    ]):
        col = 2 + ci * 2
        ws.cell(r, col, value=label).font = fn(size=9, bold=True, color=DGRAY)
        ws.cell(r, col+1, value=formula).font = fn(size=9, color=NAVY)

    # Validación: selector solo acepta 1-30
    dv = DataValidation(type="whole", operator="between",
                        formula1="1", formula2="30", allow_blank=False)
    dv.error = "Introduce un número entre 1 y 30"
    dv.errorTitle = "Número inválido"
    ws.add_data_validation(dv); dv.sqref = "C8"

    def seccion(row, titulo, color):
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"B{row}:{LAST_COL_L}{row}")
        c = ws.cell(row, 2, value=titulo)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(color); c.alignment = al(v="center")

    def fila_dato(row, label, *formulas, label_bg=LGRAY, dato_bg=WHITE):
        ws.row_dimensions[row].height = 18
        c = ws.cell(row, 2, value=label)
        c.font = fn(size=9, bold=True, color=NAVY)
        c.fill = f(label_bg); c.alignment = al(v="center"); c.border = thin_border()
        for ci, formula in enumerate(formulas):
            c = ws.cell(row, 3+ci, value=formula)
            c.font = fn(size=9, color=NAVY)
            c.fill = f(dato_bg); c.alignment = al(h="center", v="center")
            c.border = thin_border(); c.number_format = "0.00"

    # ── Secciones de evaluación (DINÁMICO — no hardcodeado) ──────────────────
    # Reg.Notas: alumno i en fila 10+i → fórmula INDEX con C$8
    evs_sorted = sorted(EVAL_RAS.keys())

    r = 11
    for ev in evs_sorted:
        ra_ids_ev = EVAL_RAS[ev]
        acts_ev   = [a for a in ACTIVIDADES if a["eval"] == ev]
        ev_color  = EV_COLORS.get(ev, MED)
        emoji     = EV_EMOJIS.get(ev, "📝")
        ra_pond   = sum(RA_BY_ID[rid]["pond"] for rid in ra_ids_ev)

        # Columna del examen en Reg.Notas
        ev_idx    = evs_sorted.index(ev)
        ex_col_l  = get_column_letter(_RN_EX_START + ev_idx)

        # Columna de cada actividad de esta eval en Reg.Notas
        act_col_ls = [get_column_letter(_RN_ACT_START + ACTIVIDADES.index(a))
                      for a in acts_ev]

        # Columna de cada RA de esta eval en Reg.Notas
        ra_col_ls = [RN_RA_COL[rid] for rid in ra_ids_ev]

        # Columnas de NOTA EVAL y RESULTADO en hoja Evaluación N
        ev_nota_l = _EV_NOTA_COL[ev]
        ev_res_l  = _EV_RES_COL[ev]

        # Cabecera de sección
        ra_ids_str = " + ".join(ra_ids_ev)
        seccion(r, f"{emoji}  EVALUACIÓN {ev}  ·  {ra_ids_str}  ({ra_pond}% del módulo)",
                ev_color)
        r += 1

        # Sub-cabecera: EXn | actividades | notas RA | NOTA EVAL
        sub_labels = [f"EX{ev} (75%)"]
        sub_labels += [f"{a['id']}·{a['ra']}" for a in acts_ev]
        sub_labels += [f"Nota {rid}" for rid in ra_ids_ev]
        sub_labels += [f"NOTA EVAL {ev}"]

        ws.row_dimensions[r].height = 16
        n_act  = len(acts_ev)
        n_ra   = len(ra_ids_ev)
        for ci, label in enumerate(sub_labels):
            is_ra   = 1 + n_act <= ci < 1 + n_act + n_ra
            is_nota = ci == len(sub_labels) - 1
            is_dual = is_ra and ra_ids_ev[ci - (1 + n_act)] == DUAL_RA
            if is_dual:
                bg = "1F6B3A"
            elif is_ra or is_nota:
                bg = ev_color
            else:
                bg = NAVY
            c = ws.cell(r, 2+ci, value=label)
            c.font = fn(size=8, bold=True, color=WHITE)
            c.fill = f(bg)
            c.alignment = al(h="center", v="center"); c.border = thin_border(WHITE)
        r += 1

        # Fila de calificaciones
        formulas  = [f"=IFERROR(INDEX({_sr('Reg. Notas')}!{ex_col_l}:{ex_col_l},10+C$8),\"\")"]
        formulas += [f"=IFERROR(INDEX({_sr('Reg. Notas')}!{cl}:{cl},10+C$8),\"\")"
                     for cl in act_col_ls]
        formulas += [f"=IFERROR(INDEX({_sr('Reg. Notas')}!{cl}:{cl},10+C$8),\"\")"
                     for cl in ra_col_ls]
        fila_dato(r, "Calificación", *formulas, dato_bg=ICE)
        r += 1

        # Fila de nota evaluación + resultado
        ws.row_dimensions[r].height = 20
        nota_col_i = 2 + 1 + n_act + n_ra   # columna donde cae NOTA EVAL
        ws.merge_cells(f"B{r}:{get_column_letter(nota_col_i - 1)}{r}")
        ws.cell(r, 2, value=f"Nota Evaluación {ev}:").font = fn(
            size=9, bold=True, color=ev_color)
        nota_c = ws.cell(r, nota_col_i,
            value=f"=IFERROR(INDEX({_sr(f'Evaluación {ev}')}!{ev_nota_l}:{ev_nota_l},10+C$8),\"\")")
        nota_c.font = fn(size=11, bold=True, color=WHITE)
        nota_c.fill = f(ev_color); nota_c.alignment = al(h="center", v="center")
        nota_c.border = thin_border(); nota_c.number_format = "0.00"
        res_c = ws.cell(r, nota_col_i + 1,
            value=f"=IFERROR(INDEX({_sr(f'Evaluación {ev}')}!{ev_res_l}:{ev_res_l},10+C$8),\"\")")
        res_c.font = fn(size=9, bold=True, color=NAVY)
        res_c.fill = f(ICE); res_c.alignment = al(h="center", v="center")
        res_c.border = thin_border()

        r += 2   # espacio entre evaluaciones

    # ── NOTA FINAL ────────────────────────────────────────────────────────────
    seccion(r, "🎓  CALIFICACIÓN FINAL DEL MÓDULO", NAVY)
    r += 1
    ws.row_dimensions[r].height = 32
    for j, ra in enumerate(RAS):
        ra_color, _ = RA_COLORS[ra["id"]]
        c = ws.cell(r, 2+j, value=f'{ra["id"]}\n({ra["pond"]}%)')
        c.font = fn(size=8, bold=True, color=WHITE)
        c.fill = f(ra_color); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)
    nf_col_i  = 2 + len(RAS)
    res_col_i = nf_col_i + 1
    for ci, label in [(nf_col_i, "NOTA FINAL"), (res_col_i, "RESULTADO")]:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(NAVY); c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)

    r += 1
    ws.row_dimensions[r].height = 26
    for j, ra in enumerate(RAS):
        _, ra_lite = RA_COLORS[ra["id"]]
        ra_col_1ord = get_column_letter(4 + j)    # D=RA1 … K=RA8 en 1ªORD
        c = ws.cell(r, 2+j,
            value=f"=IFERROR(INDEX({_sr('1ªORD')}!{ra_col_1ord}:{ra_col_1ord},10+C$8),\"\")")
        c.fill = f(ra_lite); c.border = thin_border()
        c.alignment = al(h="center", v="center"); c.number_format = "0.00"
        c.font = fn(size=10, bold=True, color=NAVY)
    nota_f = ws.cell(r, nf_col_i,
        value=f"=IFERROR(INDEX({_sr('1ªORD')}!{_ORD1_FINAL_COL_L}:{_ORD1_FINAL_COL_L},10+C$8),\"\")")
    nota_f.font = fn(size=14, bold=True, color=WHITE)
    nota_f.fill = f(NAVY); nota_f.alignment = al(h="center", v="center")
    nota_f.border = thin_border(); nota_f.number_format = "0.00"
    res_f = ws.cell(r, res_col_i,
        value=f"=IFERROR(INDEX({_sr('1ªORD')}!{_ORD1_RES_COL_L}:{_ORD1_RES_COL_L},10+C$8),\"\")")
    res_f.font = fn(size=11, bold=True, color=NAVY)
    res_f.fill = f(ICE); res_f.alignment = al(h="center", v="center")
    res_f.border = thin_border()

    # ── Área de impresión ────────────────────────────────────────────────────
    ws.print_area = f"A1:{LAST_COL_L}{r + 2}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: INFORME GRUPO — Sprint 8
# ═══════════════════════════════════════════════════════════════════════════════
def build_informe_grupo(wb):
    """Informe de grupo: estadísticas globales, semáforo de riesgo y lista completa."""

    # 1ªORD layout: filas 11-40 (alumnos 1-30), cols D…K=RAs, L=nota final, M=resultado
    ORD1_DATA_ROW_START = 11
    ORD1_DATA_ROW_END   = 40
    ORD1_NF_L  = _ORD1_FINAL_COL_L
    ORD1_RES_L = _ORD1_RES_COL_L
    N_ALUMNOS  = 30

    ws = wb.create_sheet(_sn("Informe Grupo"))
    apply_standard_header(
        ws,
        f"📈 Informe de Grupo — {MODULO['abrev']} {MODULO['curso']}",
        f"Estadísticas y seguimiento del grupo · {MODULO['nombre']} · {MODULO['anno']}",
        _sn("Informe Grupo"),
    )
    set_col_width(ws, {
        "A": "3", "B": "5",  "C": "30", "D": "12", "E": "12",
        "F": "12","G": "12", "H": "12", "I": "12", "J": "12",
        "K": "12","L": "12", "M": "14", "N": "14",
    })

    LAST_COL = get_column_letter(4 + len(RAS) + 2)  # hasta RESULTADO

    # ── KPIs del grupo ───────────────────────────────────────────────────────
    r = 8
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"B{r}:{LAST_COL}{r}")
    c = ws.cell(r, 2, value="📊  ESTADÍSTICAS DEL GRUPO")
    c.font = fn(size=11, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    kpi_rows = [
        ("Alumnos registrados", f"=COUNTA({_sr('Alumnos')}!C10:C39)"),
        ("Activos",             f"=COUNTIF({_sr('Alumnos')}!J10:J39,\"Activo\")"),
        ("Aprobados en 1ªORD (≥5)",
         f"=COUNTIF({_sr('1ªORD')}!{ORD1_RES_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_RES_L}{ORD1_DATA_ROW_END},\"APTO\")"),
        ("Suspensos en 1ªORD (<5)",
         f"=COUNTIF({_sr('1ªORD')}!{ORD1_RES_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_RES_L}{ORD1_DATA_ROW_END},\"NO APTO\")"),
        ("Media del grupo",
         f"=IFERROR(ROUND(AVERAGEIF({_sr('1ªORD')}!{ORD1_RES_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_RES_L}{ORD1_DATA_ROW_END},\"<>\",{_sr('1ªORD')}!{ORD1_NF_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_NF_L}{ORD1_DATA_ROW_END}),2),\"\")"),
        ("Nota máxima",
         f"=IFERROR(MAX({_sr('1ªORD')}!{ORD1_NF_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_NF_L}{ORD1_DATA_ROW_END}),\"\")"),
        ("Nota mínima (activos)",
         f"=IFERROR(MINIFS({_sr('1ªORD')}!{ORD1_NF_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_NF_L}{ORD1_DATA_ROW_END},{_sr('1ªORD')}!{ORD1_NF_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_NF_L}{ORD1_DATA_ROW_END},\">0\"),\"\")"),
        ("En riesgo (<4 en nota final)",
         f"=COUNTIFS({_sr('1ªORD')}!{ORD1_NF_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_NF_L}{ORD1_DATA_ROW_END},\"<4\",{_sr('1ªORD')}!{ORD1_NF_L}{ORD1_DATA_ROW_START}:"
         f"{ORD1_NF_L}{ORD1_DATA_ROW_END},\">0\")"),
    ]
    KPI_COLORS = [NAVY, NAVY, GREEN, RED, MED, "107C10", "C00000", "E87222"]

    for i, (label, formula) in enumerate(kpi_rows):
        r += 1
        ws.row_dimensions[r].height = 20
        bg = ICE if i % 2 == 0 else WHITE
        lc = ws.cell(r, 2, value=label)
        lc.font = fn(size=10, bold=True, color=NAVY)
        lc.fill = f(bg); lc.alignment = al(v="center"); lc.border = thin_border()
        vc = ws.cell(r, 3, value=formula)
        vc.font = fn(size=12, bold=True, color=WHITE)
        vc.fill = f(KPI_COLORS[i])
        vc.alignment = al(h="center", v="center"); vc.border = thin_border()
        vc.number_format = "0.00" if i >= 4 else "0"

    # ── Medias por RA ────────────────────────────────────────────────────────
    r += 2
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"B{r}:{LAST_COL}{r}")
    c = ws.cell(r, 2, value="📊  MEDIA DEL GRUPO POR RESULTADO DE APRENDIZAJE")
    c.font = fn(size=11, bold=True, color=WHITE)
    c.fill = f(MED); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 28
    for j, ra in enumerate(RAS):
        ra_color, _ = RA_COLORS[ra["id"]]
        c = ws.cell(r, 2+j, value=f'{ra["id"]}\n({ra["pond"]}%)')
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(ra_color); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    r += 1
    ws.row_dimensions[r].height = 20
    for j, ra in enumerate(RAS):
        ra_color, ra_lite = RA_COLORS[ra["id"]]
        ra_col_1ord = get_column_letter(4 + j)
        c = ws.cell(r, 2+j,
            value=f"=IFERROR(ROUND(AVERAGEIF({_sr('1ªORD')}!{ra_col_1ord}{ORD1_DATA_ROW_START}:"
                  f"{ra_col_1ord}{ORD1_DATA_ROW_END},\">0\",{_sr('1ªORD')}!{ra_col_1ord}"
                  f"{ORD1_DATA_ROW_START}:{ra_col_1ord}{ORD1_DATA_ROW_END}),2),\"\")")
        c.font = fn(size=11, bold=True, color=NAVY)
        c.fill = f(ra_lite); c.alignment = al(h="center", v="center")
        c.border = thin_border(); c.number_format = "0.00"

    # ── Lista completa del grupo ─────────────────────────────────────────────
    r += 2
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"B{r}:{LAST_COL}{r}")
    c = ws.cell(r, 2, value="👥  LISTADO COMPLETO — NOTAS FINALES")
    c.font = fn(size=11, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    # Cabecera tabla
    r += 1
    ws.row_dimensions[r].height = 30
    for ci, (label, bg) in enumerate([("Nº", NAVY), ("Apellidos y Nombre", NAVY)] +
                                      [(ra["id"], RA_COLORS[ra["id"]][0]) for ra in RAS] +
                                      [("Nota Final", NAVY), ("Resultado", NAVY)]):
        c = ws.cell(r, 2+ci, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    hdr_row = r
    # Datos — 30 alumnos desde 1ªORD
    for i in range(1, N_ALUMNOS + 1):
        r += 1
        bg = ICE if i % 2 == 0 else WHITE
        ord1_row = 10 + i
        ws.row_dimensions[r].height = 18

        # Nº
        c = ws.cell(r, 2, value=i)
        c.font = fn(size=9, bold=True, color=DGRAY)
        c.fill = f(LGRAY); c.alignment = al(h="center", v="center"); c.border = thin_border()

        # Nombre desde 1ªORD col C
        c = ws.cell(r, 3,
            value=f"=IFERROR(INDEX({_sr('1ªORD')}!C:C,{ord1_row}),\"\")")
        c.font = fn(size=9, color=NAVY)
        c.fill = f(ICE); c.alignment = al(v="center"); c.border = thin_border()

        # Nota por RA
        for j, ra in enumerate(RAS):
            _, ra_lite = RA_COLORS[ra["id"]]
            ra_col_l = get_column_letter(4 + j)
            val = ws.cell(r, 4+j,
                value=f"=IFERROR(INDEX({_sr('1ªORD')}!{ra_col_l}:{ra_col_l},{ord1_row}),\"\")")
            val.font = fn(size=9, color=NAVY)
            val.fill = f(ra_lite); val.alignment = al(h="center", v="center")
            val.border = thin_border(); val.number_format = "0.00"

        # Nota final
        nf_col_i = 4 + len(RAS)
        nf = ws.cell(r, nf_col_i,
            value=f"=IFERROR(INDEX({_sr('1ªORD')}!{ORD1_NF_L}:{ORD1_NF_L},{ord1_row}),\"\")")
        nf.font = fn(size=10, bold=True, color=NAVY)
        nf.fill = f(bg); nf.alignment = al(h="center", v="center")
        nf.border = thin_border(); nf.number_format = "0.00"

        # Resultado con semáforo
        res_col_i = nf_col_i + 1
        res = ws.cell(r, res_col_i,
            value=f"=IFERROR(INDEX({_sr('1ªORD')}!{ORD1_RES_L}:{ORD1_RES_L},{ord1_row}),\"\")")
        res.font = fn(size=9, bold=True, color=NAVY)
        res.fill = f(bg); res.alignment = al(h="center", v="center")
        res.border = thin_border()

    # ── Nota al pie + área de impresión ─────────────────────────────────────
    r += 2
    ws.merge_cells(f"B{r}:{LAST_COL}{r}")
    c = ws.cell(r, 2,
        value=f"Generado automáticamente desde 1ªORD · {MODULO['nombre']} · {MODULO['anno']}")
    c.font = fn(size=8, italic=True, color=DGRAY)
    c.fill = f(LGRAY); c.alignment = al(v="center")
    ws.row_dimensions[r].height = 14

    ws.print_area   = f"A1:{LAST_COL}{r}"
    ws.freeze_panes = "D{hdr_row_plus1}".replace("{hdr_row_plus1}", str(hdr_row + 1))
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: RÚBRICAS — Sprint 7
# ═══════════════════════════════════════════════════════════════════════════════
def build_rubricas(wb):
    """Una rúbrica por RA: CEs como indicadores × 4 niveles de desempeño.
    Descriptores auto-generados desde los datos del módulo (editables por el profesor)."""

    LEVELS = [
        # (etiqueta,       nombre,           rango,   color_header, color_fondo)
        ("Nivel 1", "No Alcanzado",  "0 – 4",  "C00000", "FDECEA"),
        ("Nivel 2", "En Proceso",    "5 – 6",  "E87222", "FDE9D7"),
        ("Nivel 3", "Alcanzado",     "7 – 8",  "4472C4", "DDEEFF"),
        ("Nivel 4", "Sobresaliente", "9 – 10", "107C10", "E2EFDA"),
    ]

    ws = wb.create_sheet(_sn("Rúbricas"))
    apply_standard_header(
        ws,
        f"🎯 Rúbricas de Evaluación — {MODULO['abrev']} {MODULO['curso']}",
        f"Indicadores de logro por Resultado de Aprendizaje · {MODULO['nombre']} · {MODULO['anno']}",
        _sn("Rúbricas"),
    )
    set_col_width(ws, {
        "A": "3",
        "B": "8",    # CE id
        "C": "30",   # Nivel 1 descriptor
        "D": "30",   # Nivel 2 descriptor
        "E": "30",   # Nivel 3 descriptor
        "F": "30",   # Nivel 4 descriptor
        "G": "9",    # Peso CE en RA
    })

    r = 8

    # ── Instrucciones ────────────────────────────────────────────────────────
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f"B{r}:G{r}")
    c = ws.cell(r, 2, value=(
        "ℹ️  Rúbricas generadas automáticamente desde la Programación. "
        "Los descriptores son orientativos — edítalos para adaptarlos a tu módulo y actividades. "
        "Peso CE: distribución equitativa por defecto (modifica si lo necesitas)."
    ))
    c.font = fn(size=9, italic=True, color=DGRAY)
    c.fill = f(ICE); c.alignment = al(v="center", wrap=True)

    # ── Una rúbrica por RA ───────────────────────────────────────────────────
    for ra in RAS:
        ces = CES_POR_RA.get(ra["id"], [])
        if not ces:
            continue

        ra_color, ra_lite = RA_COLORS[ra["id"]]
        is_dual   = (ra["id"] == DUAL_RA)
        ev        = RA_EVAL_MAP.get(ra["id"], "")
        bg_header = "1F6B3A" if is_dual else ra_color
        prefix    = "🏢 DUAL · " if is_dual else ""
        peso_ce   = round(100 / len(ces), 1)   # peso equitativo

        # ── Cabecera del RA ──────────────────────────────────────────────────
        r += 1
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"B{r}:G{r}")
        c = ws.cell(r, 2,
                    value=f"{prefix}{ra['id']} · EV{ev} · Pond. {ra['pond']}%  —  {ra['nombre']}")
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(bg_header)
        c.alignment = al(v="center", wrap=True)

        # ── Sub-cabecera: CE | Nivel 1 | Nivel 2 | Nivel 3 | Nivel 4 | Peso ──
        r += 1
        ws.row_dimensions[r].height = 36

        c = ws.cell(r, 2, value="CE / Indicador")
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(NAVY); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

        for i, (nivel, nombre, rango, hdr_color, _) in enumerate(LEVELS):
            c = ws.cell(r, 3 + i, value=f"{nivel}\n{nombre}\n({rango})")
            c.font = fn(size=9, bold=True, color=WHITE)
            c.fill = f(hdr_color)
            c.alignment = al(h="center", v="center", wrap=True)
            c.border = thin_border(WHITE)

        c = ws.cell(r, 7, value="Peso\nen RA")
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(NAVY); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

        # ── Filas de CEs (indicadores) ───────────────────────────────────────
        ra_nombre_corto = ra["nombre"].split(",")[0][:45]
        for j, ce in enumerate(ces):
            r += 1
            ws.row_dimensions[r].height = 52
            bg = ra_lite if (j % 2 == 0) else WHITE

            # CE id
            c = ws.cell(r, 2, value=ce)
            c.font = fn(size=9, bold=True, color=NAVY)
            c.fill = f(LGRAY)
            c.alignment = al(h="center", v="center")
            c.border = thin_border()

            # Descriptores auto-generados (4 niveles)
            descriptors = [
                f"No identifica {ce} ni lo aplica en el contexto de {ra_nombre_corto}.",
                f"Identifica {ce} con ayuda; lo aplica con errores o de forma incompleta.",
                f"Aplica {ce} de forma correcta y autónoma en situaciones habituales.",
                f"Aplica {ce} con criterio propio en situaciones nuevas; propone mejoras.",
            ]
            for k, (_, _, _, _, fondo) in enumerate(LEVELS):
                c = ws.cell(r, 3 + k, value=descriptors[k])
                c.font = fn(size=8, color=DGRAY, italic=True)
                c.fill = f(fondo)
                c.alignment = al(h="left", v="top", wrap=True)
                c.border = thin_border()

            # Peso CE en el RA
            c = ws.cell(r, 7, value=f"{peso_ce}%")
            c.font = fn(size=9, color=NAVY)
            c.fill = f(bg)
            c.alignment = al(h="center", v="center")
            c.border = thin_border()

        # ── Fila resumen: % total y nota de corte ────────────────────────────
        r += 1
        ws.row_dimensions[r].height = 16
        ws.merge_cells(f"B{r}:F{r}")
        c = ws.cell(r, 2,
                    value=f"Total: {len(ces)} indicadores · "
                          f"Nota RA: media ponderada de niveles × 2.5  "
                          f"(Nivel 1=0-4 · Nivel 2=5-6 · Nivel 3=7-8 · Nivel 4=9-10)")
        c.font = fn(size=8, italic=True, color=DGRAY)
        c.fill = f(LGRAY); c.alignment = al(v="center")

        c = ws.cell(r, 7, value="100%")
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg_header); c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)

        # ── Separador dorado entre RAs ───────────────────────────────────────
        r += 1
        ws.row_dimensions[r].height = 6
        for ci in range(2, 8):
            ws.cell(r, ci).fill = f(GOLD)

    ws.freeze_panes = "B9"
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: 1ª ORDINARIA — Resultado final del módulo
# ═══════════════════════════════════════════════════════════════════════════════
def build_primera_ord(wb):
    ws = wb.create_sheet(_sn("1ªORD"))
    apply_standard_header(
        ws,
        "🎓 Calificación Final — 1ª Convocatoria Ordinaria",
        f"Nota final ponderada por Resultados de Aprendizaje · "
        f"{MODULO['abrev']} {MODULO['curso']} · {MODULO['anno']}",
        _sn("1ªORD"),
    )

    # Columnas dinámicas: B=Nº, C=Nombre, D..K=RA1..RA8, L=Nota Final, M=Resultado, N=Obs
    # Col index base para RAs: columna D = índice 4
    RA_START_COL = 4   # D

    # Anchos de columna
    widths = {"A":"3","B":"5","C":"32"}
    for i, ra in enumerate(RAS):
        widths[get_column_letter(RA_START_COL + i)] = "9"
    final_col_l  = get_column_letter(RA_START_COL + len(RAS))      # L
    result_col_l = get_column_letter(RA_START_COL + len(RAS) + 1)  # M
    obs_col_l    = get_column_letter(RA_START_COL + len(RAS) + 2)  # N
    widths[final_col_l]  = "12"
    widths[result_col_l] = "12"
    widths[obs_col_l]    = "20"
    set_col_width(ws, widths)

    EVAL_COLORS_ORD = {1: "4472C4", 2: "E87222", 3: "70AD47"}

    # Fila 8: fila de evaluación por RA (badge EV1/EV2/EV3 sobre cada columna RA)
    r = 8
    ws.row_dimensions[r].height = 14
    ws.cell(r, 2, value="Evaluación:").font = fn(size=8, bold=True, color=DGRAY)
    for i, ra in enumerate(RAS):
        ev = RA_EVAL_MAP.get(ra["id"], 0)
        ev_col = EVAL_COLORS_ORD.get(ev, MED)
        c = ws.cell(r, RA_START_COL + i, value=f"EV{ev}")
        c.font = fn(size=8, bold=True, color=WHITE)
        c.fill = f(ev_col)
        c.alignment = al(h="center", v="center")

    # Fila 9: leyenda ponderaciones + nota mínima
    ws.row_dimensions[9].height = 14
    ws.cell(9, 2, value="Ponderación:").font = fn(size=8, bold=True, color=NAVY)
    for i, ra in enumerate(RAS):
        c = ws.cell(9, RA_START_COL + i, value=f'{ra["pond"]}%')
        c.font = fn(size=8, color=RA_COLORS[ra["id"]][0])
        c.alignment = al(h="center")
    ws.cell(9, ord(final_col_l)-64,
            value="≥5 → APTO").font = fn(size=8, italic=True, color=RED)

    # Fila 10: cabecera tabla
    r = 10
    ws.row_dimensions[r].height = 36

    # Nº y Nombre
    for ci, label in [(2,"Nº"),(3,"Apellidos y Nombre")]:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(NAVY); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # Una columna por RA con su ponderación (RA dual marcado en verde empresa)
    for i, ra in enumerate(RAS):
        ra_color, ra_lite = RA_COLORS[ra["id"]]
        is_dual = (ra["id"] == DUAL_RA)
        label = (f'🏢 {ra["id"]}\n({ra["pond"]}%)\nEMPRESA'
                 if is_dual else f'{ra["id"]}\n({ra["pond"]}%)')
        bg_col = "1F6B3A" if is_dual else ra_color
        c = ws.cell(r, RA_START_COL + i, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg_col)
        c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    # Nota Final, Resultado, Observaciones
    for col_l, label, bg in [
        (final_col_l,  "Nota Final",    NAVY),
        (result_col_l, "Resultado",     NAVY),
        (obs_col_l,    "Observaciones", NAVY),
    ]:
        c = ws.cell(r, ord(col_l)-64, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    ws.freeze_panes = f"D{r+1}"

    # Construir fórmula ponderada una vez
    # Nota Final = ROUND( RA1*0.13 + RA2*0.19 + ... + RA8*0.05 , 2)
    def nota_final_formula(row_n):
        parts = []
        for i, ra in enumerate(RAS):
            col_l = get_column_letter(RA_START_COL + i)
            parts.append(f"{col_l}{row_n}*{ra['pond']/100}")
        return f'=IFERROR(ROUND({"+".join(parts)},2),"")'

    # Filas de alumnos (30 alumnos, filas 11–40)
    for i in range(1, 31):
        r += 1
        bg = ICE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20

        # Nº
        c = ws.cell(r, 2, value=i)
        c.font = fn(size=10, bold=True, color=DGRAY)
        c.fill = f(LGRAY); c.alignment = al(h="center", v="center"); c.border = thin_border()

        # Nombre — auto desde Alumnos
        c = ws.cell(r, 3, value=nombre_formula(r))
        c.fill = f(ICE); c.border = thin_border(); c.alignment = al(v="center")
        c.font = fn(size=9, color=NAVY)

        # Notas por RA — AUTO desde Reg. Notas (cascada cerrada)
        # Sprint 6: sustituye entrada manual por fórmula =IFERROR({_sr('Reg. Notas')}!XX{row},"")
        for j, ra in enumerate(RAS):
            _, ra_lite = RA_COLORS[ra["id"]]
            is_dual = (ra["id"] == DUAL_RA)
            cell_bg = "C8E6C9" if is_dual else ra_lite
            rn_col  = RN_RA_COL[ra["id"]]
            c = ws.cell(r, RA_START_COL + j,
                        value=f"=IFERROR({_sr('Reg. Notas')}!{rn_col}{r},\"\")")
            c.fill = f(cell_bg); c.border = thin_border()
            c.alignment = al(h="center", v="center")
            c.number_format = "0.00"
            c.font = fn(size=9, color="1F6B3A" if is_dual else NAVY)

        # Nota Final ponderada
        final_ci = ord(final_col_l) - 64
        c = ws.cell(r, final_ci, value=nota_final_formula(r))
        c.font = fn(size=10, bold=True, color=NAVY)
        c.fill = f(ICE); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.number_format = "0.00"

        # Resultado APTO / NO APTO
        result_ci = ord(result_col_l) - 64
        nf_ref = f"{final_col_l}{r}"
        c = ws.cell(r, result_ci,
                    value=f'=IF({nf_ref}="","",IF({nf_ref}>=5,"APTO","NO APTO"))')
        c.font = fn(size=10, bold=True, color=GREEN)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")

        # Observaciones
        obs_ci = ord(obs_col_l) - 64
        c = ws.cell(r, obs_ci)
        c.fill = f(bg); c.border = thin_border()

    # Fila resumen del grupo
    r += 2
    ws.row_dimensions[r].height = 22
    final_ci   = ord(final_col_l)  - 64
    result_ci  = ord(result_col_l) - 64

    for ci, label in [(2,"GRUPO"),(3,"")]:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(NAVY); c.border = thin_border(WHITE)
        c.alignment = al(h="center", v="center")

    # Media de cada RA
    for j, ra in enumerate(RAS):
        ra_color, _ = RA_COLORS[ra["id"]]
        col_l = get_column_letter(RA_START_COL + j)
        c = ws.cell(r, RA_START_COL + j,
                    value=f'=IFERROR(ROUND(AVERAGE({col_l}11:{col_l}40),2),"")')
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(ra_color); c.border = thin_border(WHITE)
        c.alignment = al(h="center", v="center")

    # Media final del grupo
    fl = final_col_l
    c = ws.cell(r, final_ci,
                value=f'=IFERROR(ROUND(AVERAGE({fl}11:{fl}40),2),"")')
    c.font = fn(size=10, bold=True, color=WHITE)
    c.fill = f(NAVY); c.border = thin_border(WHITE)
    c.alignment = al(h="center", v="center")

    # % aprobados
    rl = result_col_l
    c = ws.cell(r, result_ci,
                value=f'=IFERROR(TEXT(COUNTIF({rl}11:{rl}40,"APTO")/COUNTA({rl}11:{rl}40),"0%"),"")')
    c.font = fn(size=10, bold=True, color=WHITE)
    c.fill = f(NAVY); c.border = thin_border(WHITE)
    c.alignment = al(h="center", v="center")

    # ── Nota de evaluación dual ──────────────────────────────────────────────
    r += 2
    last_col = ord(obs_col_l) - 64
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:{obs_col_l}{r}")
    c = ws.cell(r, 2, value="🏢  EVALUACIÓN DUAL — Instrucciones y Justificación")
    c.font = fn(size=10, bold=True, color=WHITE)
    c.fill = f("1F6B3A"); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 90
    ws.merge_cells(f"B{r}:{obs_col_l}{r}")
    c = ws.cell(r, 2, value=DUAL_JUSTIFICACION)
    c.font = fn(size=9, italic=True, color=DGRAY)
    c.fill = f("E8F5E9")
    c.alignment = al(h="left", v="top", wrap=True)
    c.border = thin_border("1F6B3A")

    r += 1
    ws.row_dimensions[r].height = 16
    ws.merge_cells(f"B{r}:{obs_col_l}{r}")
    dual_ra_col = get_column_letter(RA_START_COL + next(
        i for i, ra in enumerate(RAS) if ra["id"] == DUAL_RA))
    c = ws.cell(r, 2,
                value=f"→ Introduce la nota de la empresa en la columna {DUAL_RA} "
                      f"(columna {dual_ra_col}, fondo verde).  "
                      f"Peso en nota final: {DUAL_PCT_NOTA*100:.0f}% = {DUAL_PCT_NOTA} puntos.  "
                      f"Base legal: Ley Orgánica 3/2022, art. 42.")
    c.font = fn(size=9, bold=True, color="1F6B3A")
    c.fill = f("E8F5E9"); c.alignment = al(v="center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: 2ª ORDINARIA — Solo alumnos no superados en 1ªORD
# ═══════════════════════════════════════════════════════════════════════════════
def build_segunda_ord(wb):
    # Columna de nota final en 1ªORD (referencia cruzada)
    _ORD1_RA_START  = 4                          # col D
    _ORD1_FINAL_COL = _ORD1_RA_START + len(RAS)  # col L (si hay 8 RAs)
    _ORD1_FINAL_L   = get_column_letter(_ORD1_FINAL_COL)

    ws = wb.create_sheet(_sn("2ªORD"))
    apply_standard_header(
        ws,
        "📋 Calificación Final — 2ª Convocatoria Extraordinaria",
        f"Solo alumnado NO APTO en 1ª Convocatoria · "
        f"{MODULO['abrev']} {MODULO['curso']} · {MODULO['anno']}",
        _sn("2ªORD"),
    )
    set_col_width(ws, {"A":"3","B":"5","C":"35","D":"14","E":"14",
                        "F":"14","G":"12","H":"20"})

    # Nota informativa
    r = 8
    ws.row_dimensions[r].height = 20
    ws.cell(r, 2,
            value=(
                "ℹ️  Registra únicamente el alumnado con resultado NO APTO en 1ª Convocatoria. "
                "Introduce el número de alumno/a (columna B) para que el nombre y la nota de 1ªORD "
                "se auto-rellenen. La nota extraordinaria sustituye la nota final del módulo."
            )).font = fn(size=9, italic=True, color=DGRAY)

    # Fila 9: leyenda columnas
    ws.row_dimensions[9].height = 12
    ws.cell(9, 4, value=f"← Auto desde 1ªORD col {_ORD1_FINAL_L}").font = fn(size=8, italic=True, color=DGRAY)

    # Cabecera fila 10
    r = 10
    ws.row_dimensions[r].height = 32
    headers = [
        (2, "Nº\nalumno",         NAVY),
        (3, "Apellidos y Nombre", NAVY),
        (4, "Nota 1ªORD",        RED),
        (5, "Nota 2ªORD\n(Extraordinaria)", ORNG),
        (6, "Nota Final\n(computa)", NAVY),
        (7, "Resultado",          NAVY),
        (8, "Observaciones",      NAVY),
    ]
    for col_i, label, bg in headers:
        c = ws.cell(r, col_i, value=label)
        c.font = fn(size=10, bold=True, color=WHITE)
        c.fill = f(bg)
        c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    ws.freeze_panes = f"D{r+1}"

    for i in range(1, 21):   # Máx 20 alumnos en extraordinaria
        r += 1
        bg = ICE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 20

        # Col B: número de alumno (entrada manual — referencia a Alumnos y 1ªORD)
        c = ws.cell(r, 2)
        c.font = fn(size=10, bold=True, color="C00000")
        c.fill = f("FDECEA"); c.alignment = al(h="center", v="center"); c.border = thin_border()
        c.number_format = "0"

        # Col C: Nombre — auto desde Alumnos via número de alumno en B
        # Alumnos: cabecera fila 9, datos filas 10-39 → alumno n en fila 9+n
        b_ref = f"B{r}"
        c = ws.cell(r, 3,
                    value=f"=IFERROR(IF({b_ref}=\"\",\"\","
                          f"INDEX({_sr('Alumnos')}!C:C,9+{b_ref})"
                          f"&\", \"&INDEX({_sr('Alumnos')}!D:D,9+{b_ref})),\"\")")
        c.fill = f(ICE); c.border = thin_border(); c.alignment = al(v="center")
        c.font = fn(size=9, color=NAVY)

        # Col D: Nota 1ªORD — auto desde 1ªORD via número de alumno
        # 1ªORD: datos en filas 11-40 → alumno n en fila 10+n
        c = ws.cell(r, 4,
                    value=f"=IFERROR(IF({b_ref}=\"\",\"\","
                          f"INDEX({_sr('1ªORD')}!{_ORD1_FINAL_L}:{_ORD1_FINAL_L},10+{b_ref})),\"\")")
        c.fill = f("F5DCDC"); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.number_format = "0.00"
        c.font = fn(size=9, color=RED)

        # Col E: Nota 2ªORD (examen extraordinario — entrada manual)
        c = ws.cell(r, 5)
        c.fill = f("FDE9D7"); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.number_format = "0.00"

        # Col F: Nota Final que computa (2ªORD si existe, si no 1ªORD)
        d_ref = f"D{r}"; e_ref = f"E{r}"
        c = ws.cell(r, 6,
                    value=f'=IF({e_ref}="",IF({d_ref}="","",{d_ref}),{e_ref})')
        c.font = fn(size=10, bold=True, color=NAVY)
        c.fill = f(ICE); c.border = thin_border()
        c.alignment = al(h="center", v="center")
        c.number_format = "0.00"

        # Col G: Resultado APTO / NO APTO
        f_ref = f"F{r}"
        c = ws.cell(r, 7,
                    value=f'=IF({f_ref}="","",IF({f_ref}>=5,"APTO","NO APTO"))')
        c.font = fn(size=10, bold=True, color=GREEN)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")

        # Col H: Observaciones
        c = ws.cell(r, 8)
        c.fill = f(bg); c.border = thin_border()

    # Nota aclaratoria al pie
    r += 2
    ws.row_dimensions[r].height = 16
    ws.cell(r, 2,
            value="★  Introduce el número de alumno/a (1-30) en columna B. "
                  "El nombre y la nota de 1ªORD se cargarán automáticamente. "
                  "El alumnado APTO en 1ª convocatoria no debe aparecer aquí.").font = fn(
        size=9, italic=True, color=NAVY)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA: DASHBOARD — Cuadro de Mando
# ═══════════════════════════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.create_sheet(_sn("Dashboard"))
    apply_standard_header(
        ws,
        "📊 Cuadro de Mando — Dashboard",
        f"{MODULO['nombre']}  ·  {MODULO['curso']}  ·  {MODULO['anno']}",
        _sn("Dashboard"),
    )
    set_col_width(ws, {
        "A": "2", "B": "5", "C": "32",
        "D": "10", "E": "10", "F": "10",
        "G": "12", "H": "16", "I": "12",
    })

    NF_L  = _ORD1_FINAL_COL_L   # col Nota Final en 1ªORD (dinámica, p.ej. "L")
    RES_L = _ORD1_RES_COL_L     # col Resultado en 1ªORD  (p.ej. "M")

    # ── KPI helper ──────────────────────────────────────────────────────────────
    def kpi_box(row, col, label, formula, bg):
        cl  = get_column_letter(col)
        cl2 = get_column_letter(col + 1)
        ws.row_dimensions[row].height = 13
        ws.row_dimensions[row + 1].height = 30
        ws.merge_cells(f"{cl}{row}:{cl2}{row}")
        c = ws.cell(row, col, value=label)
        c.font = fn(size=8, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)
        ws.merge_cells(f"{cl}{row+1}:{cl2}{row+1}")
        c = ws.cell(row + 1, col, value=formula)
        c.font = fn(size=20, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center")
        c.border = thin_border(WHITE)

    # ── Sección 1: KPIs globales ─────────────────────────────────────────────
    r = 8
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  INDICADORES GLOBALES DEL GRUPO")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    # Fila KPI 1: Activos · Aptos · No Aptos
    ord_ref  = _sr("1ªORD")
    alum_ref = _sr("Alumnos")
    kpi_box(9, 2, "ALUMNOS ACTIVOS",
            f"=COUNTA({alum_ref}!C10:C39)",
            NAVY)
    kpi_box(9, 4, "✅  APTOS",
            f'=COUNTIF({ord_ref}!{RES_L}11:{RES_L}40,"APTO")',
            "107C10")
    kpi_box(9, 6, "❌  NO APTOS",
            f'=COUNTIF({ord_ref}!{RES_L}11:{RES_L}40,"NO APTO")',
            "C00000")

    # Spacer
    ws.row_dimensions[11].height = 5

    # Fila KPI 2: Media · Máxima · En riesgo (<4)
    kpi_box(12, 2, "MEDIA GRUPO",
            f'=IFERROR(ROUND(AVERAGE({ord_ref}!{NF_L}11:{NF_L}40),2),"")',
            MED)
    kpi_box(12, 4, "NOTA MÁXIMA",
            f'=IFERROR(MAX({ord_ref}!{NF_L}11:{NF_L}40),"–")',
            "375623")
    kpi_box(12, 6, "⚠️  EN RIESGO",
            f'=COUNTIFS({ord_ref}!{NF_L}11:{NF_L}40,">0",{ord_ref}!{NF_L}11:{NF_L}40,"<4")',
            "E87222")

    ws.row_dimensions[14].height = 5

    # ── Separador ─────────────────────────────────────────────────────────────
    r = 15
    ws.row_dimensions[r].height = 5
    for ci in range(2, 10):
        ws.cell(r, ci).fill = f(GOLD)

    # ── Sección 2: tabla semáforo por alumno ──────────────────────────────────
    r = 16
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  SEGUIMIENTO INDIVIDUAL — SEMÁFORO DE RIESGO")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    r = 17
    ws.row_dimensions[r].height = 30
    EVAL_BG = {1: "4472C4", 2: "E87222", 3: "107C10"}
    hdrs = [
        (2,  "Nº",                NAVY),
        (3,  "Apellidos y Nombre", NAVY),
        (4,  f"Eval 1\n(nota)",   EVAL_BG[1]),
        (5,  f"Eval 2\n(nota)",   EVAL_BG[2]),
        (6,  f"Eval 3\n(nota)",   EVAL_BG[3]),
        (7,  "Nota Final",        NAVY),
        (8,  "Estado",            NAVY),
        (9,  "% Evals OK",        MED),
    ]
    for ci, label, bg in hdrs:
        c = ws.cell(r, ci, value=label)
        c.font = fn(size=9, bold=True, color=WHITE)
        c.fill = f(bg); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    ws.freeze_panes = "D18"

    EVAL_LITE = {1: "DDEEFF", 2: "FDE9D7", 3: "E2EFDA"}
    for i in range(1, 31):
        r += 1
        bg = ICE if i % 2 == 0 else WHITE
        ws.row_dimensions[r].height = 18

        # Nº
        c = ws.cell(r, 2, value=i)
        c.font = fn(size=10, bold=True, color=DGRAY)
        c.fill = f(LGRAY); c.alignment = al(h="center", v="center"); c.border = thin_border()

        # Nombre
        c = ws.cell(r, 3, value=nombre_formula(r))
        c.fill = f(ICE); c.border = thin_border(); c.alignment = al(v="center")
        c.font = fn(size=9, color=NAVY)

        # Notas por evaluación (referencia directa a cada hoja Evaluación N)
        for j, ev in enumerate(sorted(EVAL_RAS.keys())):
            ev_col      = _EV_NOTA_COL[ev]
            ev_sheet_r  = _sr(f"Evaluación {ev}")
            c = ws.cell(r, 4 + j,
                        value=f"=IFERROR(INDEX({ev_sheet_r}!{ev_col}:{ev_col},10+{i}),\"\")")
            c.fill = f(EVAL_LITE[ev]); c.border = thin_border()
            c.alignment = al(h="center", v="center"); c.number_format = "0.00"
            c.font = fn(size=9, color=NAVY)

        # Nota Final (desde 1ªORD)
        c = ws.cell(r, 7,
                    value=f"=IFERROR(INDEX({_sr('1ªORD')}!{NF_L}:{NF_L},10+{i}),\"\")")
        c.font = fn(size=10, bold=True, color=NAVY)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center"); c.number_format = "0.00"

        # Estado semáforo
        nf = f"G{r}"
        c = ws.cell(r, 8,
                    value=f'=IF({nf}="","–",IF({nf}>=7,"🟢 NOTABLE",IF({nf}>=5,"🟡 APTO","🔴 NO APTO")))')
        c.font = fn(size=9, bold=True, color=DGRAY)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")

        # % evaluaciones con nota introducida
        d_ref = f"D{r}"; e_ref = f"E{r}"; ff_ref = f"F{r}"
        c = ws.cell(r, 9,
                    value=f'=IFERROR(TEXT(COUNTIF({d_ref}:{ff_ref},">0")/3,"0%"),"–")')
        c.font = fn(size=9, color=DGRAY)
        c.fill = f(bg); c.border = thin_border()
        c.alignment = al(h="center", v="center")

    # ── Separador ─────────────────────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 5
    for ci in range(2, 10):
        ws.cell(r, ci).fill = f(GOLD)

    # ── Sección 3: medias del grupo por RA ────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2, value="  RENDIMIENTO POR RA — Media del grupo (desde Reg. Notas)")
    c.font = fn(size=9, bold=True, color=WHITE)
    c.fill = f(NAVY); c.alignment = al(v="center")

    r += 1
    ws.row_dimensions[r].height = 26
    for j, ra in enumerate(RAS):
        ra_color, _ = RA_COLORS[ra["id"]]
        ev = RA_EVAL_MAP.get(ra["id"], "")
        c = ws.cell(r, 2 + j, value=f'{ra["id"]}\nEV{ev} · {ra["pond"]}%')
        c.font = fn(size=8, bold=True, color=WHITE)
        c.fill = f(ra_color); c.alignment = al(h="center", v="center", wrap=True)
        c.border = thin_border(WHITE)

    r += 1
    ws.row_dimensions[r].height = 24
    for j, ra in enumerate(RAS):
        _, ra_lite = RA_COLORS[ra["id"]]
        rn_col = RN_RA_COL[ra["id"]]
        c = ws.cell(r, 2 + j,
                    value=f'=IFERROR(ROUND(AVERAGE({_sr("Reg. Notas")}!{rn_col}11:{rn_col}40),2),"–")')
        c.font = fn(size=14, bold=True, color=NAVY)
        c.fill = f(ra_lite); c.alignment = al(h="center", v="center")
        c.border = thin_border(); c.number_format = "0.00"

    # ── Pie: nota aclaratoria ─────────────────────────────────────────────────
    r += 2
    ws.row_dimensions[r].height = 14
    ws.merge_cells(f"B{r}:I{r}")
    c = ws.cell(r, 2,
                value="★  Los datos se actualizan automáticamente al introducir notas. "
                      "Semáforo: 🟢 ≥7 · 🟡 5-6.9 · 🔴 <5.  En riesgo = nota final < 4.")
    c.font = fn(size=8, italic=True, color=DGRAY)
    c.alignment = al(v="center")

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# HOJA OCULTA: _IA_Config (Asistente IA — Sprint 2.6)
# ═══════════════════════════════════════════════════════════════════════════════
def build_ia_config(wb, modules_config):
    """Crea hoja oculta _IA_Config con instrucciones del Asistente IA."""
    ws = wb.create_sheet("_IA_Config")
    ws.sheet_state = "hidden"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    def _h(row, label, value=""):
        ws.cell(row, 1, label).font  = fn(bold=True, color=NAVY)
        ws.cell(row, 2, value).font  = fn(color="333333")

    def _r(row, value):
        c = ws.cell(row, 2, value)
        c.font = fn(name="Courier New", size=9, color="2D6A9F")

    ws.cell(1, 1, "EvalFP — Asistente IA (Sprint 2.6)").font = fn(bold=True, size=13, color=NAVY)
    ws.cell(2, 1, "Generación de contenido pedagógico asistido por IA").font = fn(italic=True, color=DGRAY)
    ws.cell(3, 1, "─" * 60)

    _h(5,  "¿Qué genera?",     "Descriptores de rúbrica · Propuestas de actividades · Informes individuales")
    _h(6,  "Script",           "scripts/ai_asistente.py")
    _h(7,  "Dependencia",      "pip install anthropic   (o pip install openai)")
    _h(8,  "API Key (Claude)", "Configura la variable de entorno: ANTHROPIC_API_KEY=sk-ant-...")
    _h(9,  "API Key (OpenAI)", "Configura la variable de entorno: OPENAI_API_KEY=sk-...")
    _h(10, "Sin API Key",      "Funciona en modo DEMO con texto de ejemplo (sin llamada a API)")

    ws.cell(12, 1, "Comandos de ejemplo").font = fn(bold=True, size=11, color=NAVY)
    ws.cell(12, 1).fill = f(ICE)
    ws.cell(12, 2).fill = f(ICE)

    _r(13, "python scripts/ai_asistente.py --ayuda")
    for i, (mod, _grp) in enumerate(modules_config):
        abrev = mod.MODULO["abrev"].lower()
        ra1   = mod.RAS[0]["id"]
        ra2   = mod.RAS[1]["id"] if len(mod.RAS) > 1 else ra1
        _r(14 + i*4,     f"# Módulo {mod.MODULO['abrev']}")
        _r(15 + i*4,     f"python scripts/ai_asistente.py rubrica   --modulo {abrev}_data --ra {ra1}")
        _r(16 + i*4,     f"python scripts/ai_asistente.py actividad --modulo {abrev}_data --ra {ra2} --n 3")
        _r(17 + i*4,     f"python scripts/ai_asistente.py todo      --modulo {abrev}_data --salida ia_output/{abrev}")

    last = 14 + len(modules_config) * 4 + 1
    ws.cell(last,   1, "Generación completa").font = fn(bold=True, color=NAVY)
    _r(last + 1,    "python scripts/build_template.py --ia   # genera xlsx + contenido IA")
    ws.cell(last + 3, 1, f"Módulos configurados: {', '.join(m.MODULO['abrev'] for m,_ in modules_config)}").font = fn(italic=True, color=DGRAY)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def _build_module_sheets(wb):
    """Genera el conjunto completo de hojas para el módulo activo en _SHEET_PREFIX."""
    build_programacion(wb)
    build_alumnos(wb)
    build_actividades(wb)
    build_reg_notas(wb)
    build_resumen(wb)
    for n in sorted(EVAL_RAS.keys()):
        build_evaluacion(wb, n)
    build_primera_ord(wb)
    build_segunda_ord(wb)
    build_rubricas(wb)
    build_informe_grupo(wb)
    build_boletin(wb)
    build_dashboard(wb)
    build_data_sheets(wb)


def main():
    # ── Flags de línea de comandos ───────────────────────────────────────────
    args         = sys.argv[1:]
    flag_ia      = "--ia" in args
    flag_apuntes = "--apuntes" in args
    flag_verbose = "--verbose" in args or "-v" in args

    # ── Cargar configuración del profesor ────────────────────────────────────
    try:
        sys.path.insert(0, os.path.dirname(__file__))
        from teacher_config import TEACHER_MODULES
    except ImportError:
        # Fallback: mono-módulo (compatibilidad con ejecución directa)
        import modules.iso_data as iso_data
        TEACHER_MODULES = [(iso_data, "Grupo A")]

    wb = Workbook()
    wb.remove(wb.active)

    # ── Hojas globales (sin prefijo) ─────────────────────────────────────────
    # Cargar primer módulo como contexto para build_inicio (datos del profesorado)
    first_mod, _ = TEACHER_MODULES[0]
    _set_module_context(first_mod, prefix="")
    build_inicio(wb)
    build_config(wb)
    build_mis_modulos(wb, TEACHER_MODULES)
    build_biblioteca(wb, TEACHER_MODULES)
    build_calendario(wb, TEACHER_MODULES)
    build_panel_diario(wb, TEACHER_MODULES)
    build_dashboard_global(wb, TEACHER_MODULES)
    build_ia_config(wb, TEACHER_MODULES)   # hoja oculta _IA_Config

    # ── Hojas de módulo (con prefijo "{ABREV} · ") ───────────────────────────
    for mod_data, grupo in TEACHER_MODULES:
        prefix = f"{mod_data.MODULO['abrev']} · "
        print(f"  ▶ {mod_data.MODULO['abrev']} — {grupo}")
        _set_module_context(mod_data, prefix=prefix)
        _build_module_sheets(wb)

    # ── Guardar xlsx ─────────────────────────────────────────────────────────
    out_dir = os.path.join(os.path.dirname(__file__), "..", "src")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "EvalFP.xlsx")
   
    for i, ws in enumerate(wb.worksheets, start=1):
        print(f"{i:02d} -> {ws.title}")
   
    wb.save(out_path)
    # Post-save XML fixes: openpyxl genera OOXML inválido en 3 puntos que Excel 365
    # detecta como "contenido dañado" y muestra el diálogo de reparación.
    #
    # Fix 1 – workbook.xml:      <workbookProtection/> vacío → eliminarlo
    # Fix 2 – styles.xml:        <patternFill/> sin patternType → patternType="none"
    # Fix 3 – workbook.xml.rels: Target="/xl/…" absoluto → relativo
    import zipfile as _zf, shutil as _sh, re as _re
    _tmp = out_path + ".tmp"
    with _zf.ZipFile(out_path, 'r') as _zin, _zf.ZipFile(_tmp, 'w', _zf.ZIP_DEFLATED) as _zout:
        for _item in _zin.infolist():
            _data = _zin.read(_item.filename)
            if _item.filename == 'xl/workbook.xml':
                _xml = _data.decode('utf-8')
                _xml = _re.sub(r'<workbookProtection[^/]*/>', '', _xml)
                _data = _xml.encode('utf-8')
            elif _item.filename == 'xl/styles.xml':
                _xml = _data.decode('utf-8')
                # Fix 2: <patternFill/> → <patternFill patternType="none"/>
                _xml = _xml.replace('<patternFill/>', '<patternFill patternType="none"/>')
                # Fix 5: openpyxl escribe colores ARGB con alpha=00 (transparente).
                # Excel requiere alpha=FF (opaco). Afecta a 190 colores en borders/fills/fonts.
                _xml = _re.sub(r'rgb="00([0-9A-Fa-f]{6})"', r'rgb="FF\1"', _xml)
                _data = _xml.encode('utf-8')
            elif _item.filename == 'xl/_rels/workbook.xml.rels':
                _xml = _data.decode('utf-8')
                _xml = _xml.replace('Target="/xl/worksheets/', 'Target="worksheets/')
                _xml = _xml.replace('Target="/xl/styles.xml"', 'Target="styles.xml"')
                _xml = _xml.replace('Target="/xl/theme/', 'Target="theme/')
                # Self-close all remaining truly empty cells
               # _xml = _xml.replace(' t="inlineStr"></c>', '></c>')
                #_xml = _re.sub(r'(<c [^>]*?)></c>', r'\1/>', _xml)
                #_data = _xml.encode('utf-8')
            _zout.writestr(_item, _data)
    _sh.move(_tmp, out_path)
    n_vis    = len([s for s in wb.sheetnames if wb[s].sheet_state == "visible"])
    n_hidden = len([s for s in wb.sheetnames if wb[s].sheet_state != "visible"])
    print(f"✅ EvalFP.xlsx — {len(TEACHER_MODULES)} módulo(s) · {n_vis} hojas visibles · {n_hidden} ocultas")
    print(f"   Guardado en: {os.path.abspath(out_path)}")

    # ── Generación de apuntes HTML (opcional, --apuntes flag) ────────────────
    if flag_apuntes:
        print("\n📄 Generando apuntes HTML…")
        try:
            from build_apuntes import generar_modulo
            from ai_asistente import IAAsistente
            ia      = IAAsistente()
            ap_base = os.path.join(os.path.dirname(__file__), "..", "apuntes")
            for mod_data, _grupo in TEACHER_MODULES:
                abrev = mod_data.MODULO["abrev"]
                print(f"  📚 {abrev}…")
                archivos = generar_modulo(mod_data, ia, __import__("pathlib").Path(ap_base))
                print(f"     ✅ {len(archivos)} apunte(s) en apuntes/04-FP/{abrev}/")
        except ImportError as e:
            print(f"  ⚠️  {e}. Asegúrate de que build_apuntes.py está en scripts/.")
        except Exception as e:
            print(f"  ⚠️  Error generando apuntes: {e}")

    # ── Generación IA (opcional, --ia flag) ──────────────────────────────────
    if flag_ia:
        print("\n🤖 Iniciando Asistente IA…")
        try:
            from ai_asistente import IAAsistente
            ia     = IAAsistente()
            ia_dir = os.path.join(os.path.dirname(__file__), "..", "ia_output")
            for mod_data, _grupo in TEACHER_MODULES:
                abrev  = mod_data.MODULO["abrev"].lower()
                subdir = os.path.join(ia_dir, abrev)
                print(f"  📚 Generando contenido para {mod_data.MODULO['abrev']}…")
                archivos = ia.generar_todo_modulo(mod_data, subdir)
                print(f"     ✅ {len(archivos)} archivos en ia_output/{abrev}/")
        except ImportError:
            print("  ⚠️  ai_asistente.py no encontrado en scripts/. Omitiendo paso IA.")
        except Exception as e:
            print(f"  ⚠️  Error en Asistente IA: {e}")

if __name__ == "__main__":
    main()
